#
# 엑셀의 시간 서식을 읽기
# 소리 사운드 OFF 하기

# 퀴즈 유무 검사 :
#    종료시간 지연하기 , 퀴즈가 끝나면 종료 타이머 시작
#      빔프로젝터 off, pc off
# OFF, DOWN 연장시 기존 스케쥴과 충돌 여부 검사하기
# 단축키로 ON, OFF, OS_OFF 기능 동작하기
#
# 시리얼 포트를 필요할 때만 열수 있도록 전송직전

'''
[주의] 
ON, OFF 동작후 프로젝터가 충분한 대기 시간이 있어야만,
다음 ON, OFF명령어를 받아드리고 동작한다. 그렇지 않으면, 그냥 무시해 버림


QSerial, Qtimer 모두 Thread를 사용한다. 충돌방생
  Thread 내에서 QSerial 는 정상동작
  Thread 내에서 Qtimer 사용시 timer_timeout.connect() 미동작
    self.exec() 하면 동작하지만, timer.stop() 하면 경고 발생
    Qtimer와  QSerial과 같이 사용 못한다. 오류 발생
-> time.sleep(초) 형태로 교체

전원 켜기 동작
POWE ON : "C00"[CR]
    Receive Successfully :  [ASK] [CR] ACKnowledg   - 06 0D
    Receive Unsuccessfully : "?" [CR]

전원 끄기 동작
POWE OFF : "C01"[CR]
    Receive Successfully :  [ASK] [CR]  - 06 0D
    Receive Unsuccessfully : "?" [CR]

전원 끄기 동작 확인 / 다시한번 수행하면 강제 전원끄기
POWE OFF : "C02"[CR]
    Receive Successfully :  [ASK] [CR]  - 06 0D
    Receive Unsuccessfully : "?" [CR]

'''

# openpyxl
# https://openpyxl.readthedocs.io/en/stable/

# https://lunikism.com/entry/통신-제어-문자STX-ETX-CR-LF


from openpyxl import Workbook, load_workbook
from setup import *
import re, datetime, time, sys, os
from PySide6.QtCore import QTime
from PySide6.QtSerialPort import QSerialPort, QSerialPortInfo

from enum import Enum

#####################################################################
#     동작 테스트용 설정
#####################################################################
# TEST = True        # 테스트용
TEST = False

# KEY_TEST = True    # 테스트용
KEY_TEST = False

# SHUTDOWN = False  # 테스트용
SHUTDOWN = True

TEST_CURRENT_TIME = QTime(0, 30)
#####################################################################


WEEKDAY_NAMES = ['월요일', '화요일', '수요일', '목요일', '금요일', '토요일', '일요일']
WEEKDAY_ENG = ['MON', 'TUS', 'WED', 'THU', 'FRI', 'SAT', 'SUN']
WEEKDAY_CELL = [0, 2, 4, 6, 8, 10, 12]

SCH_ROW_MAX = 50
CONTROL_CMD = ['ON', 'OFF']
# 정규식으로 시간 체크 https://htmlstory.tistory.com/32
#   https://seoyeonhwng.medium.com/파이썬-정규표현식-뿌시기-e4e96c834991
# TIME_PATTERN = r"^([01][0-9]|2[0-3]):([0-5][0-9])$"                 # 시간형식 체크 정규식 hh:mm
TIME_PATTERN = r"^([0-9]|[01][0-9]|2[0-3]):([0-9]|[0-5][0-9])$"       # 시간형식 체크 정규식 hh:mm


WAIT_FOR_READY_READ = 1000  # 1초 동안 ACK 대기


####################################################################################
# 파이썬 3.4 이후 열거형 지원

# 전체 스케쥴              [ 0, 1,   2    ,  3                , 4, 0 ]
# self.daySchedule_data = [[0, ON, '08:30', QTime(8, 30, 0, 0), 0, 0 ], [1, , ,], ... ]
# self.daySchedule_data = [[index, CTL, TIME, QTIME, 0, 0], [] ]

# 남은 스케쥴
# self.sch_remaining_list = 
SCH_ITEM = [None, None, None, None, 0, 0]
class Sch(Enum):
    INDEX = 0       # 스케줄 인덕스
    CTL = 1         # 제어 명령어
    TTIME = 2       # text 시간
    QTIME = 3       # Qtime 시간
    SEND = 4        # 제어 SEND 전송 횟수
    RESP = 5        # 제어 수신 횟수 0:미처리, 1:OK 수신, -1: 미수신1회, -2: 미수신 2회
    # 사용법 : Sch.CTL.value
####################################################################################


class ThreadScheduler(QThread):
    message_signal = Signal(str)
    music_signal = Signal(str)
    # daychange_signal = Signal(bool)

    def __init__(self):
        super().__init__()

        self.sch_remaining_list = []     # 남은 스케줄 저장
        self.sch_shutdown_timer = False  # 종료 타이머
        self.sch_prev = None        # 이전 스케줄 로그
        # self._run()      # 쓰레드에서느 불필요 [테스트]

        # self.sch_quiz_progress_set = True # False # 퀴즈 진행중일 때 SCH_QUIZ_PROGRES_DELAY_MIN
        self.sch_quiz_progress_set = False

    def stop(self):
        """Sets run flag to False and waits for thread to finish"""
        # self._run_flag = False
        self.quit()
        self.wait(5000)


    def run(self):
        # self.message_signal.emit('> 1 Serial Thread run... ')

        # self.daySchedule_data  -> save
        self.schedule_Int_Load()

        # self.message_signal.emit('> 2 Serial run... ')
        # self._run_flag = self.serial_Init()
        # self.message_signal.emit(str(self._run_flag) )

        # [test]
        # self.timer_timeout()


        ###########################################################################
        # self.timer = QTimer()
        # self.timer.setInterval(TIMER_SCH_INTERVER)       # TIMER_SCH_INTERVER
        # self.timer.timeout.connect(self.timer_timeout)

        # self.timer.start()
        # print(' * timer_start ...')serial_InitW

        # self.sch_timer_chk_cnt = SCH_TIMER_CHK_CNT

        # [중요] Qthread 에서 Qtimer 를 중지시킬수 없다.
        # self.timer.stop()     # [경고 발생]

        # [중요] Qthread 안에서 Qtimer 사용하기
        # pyqt How to use Qtimer on Qthread?
        #   https://stackoverflow.com/questions/52036021/qtimer-on-a-qthread
        # self.exec()

        # NEW

        # 시리얼 초기 검사
        print('*******************************')
        print('**      시리얼 포트 Check    **')
        print('*******************************')
        ret_serial_init = self.serial_Init()
        if  ret_serial_init:
            self.serial_Stop()
            print(f'시리얼 {self.proj_port_name} Port 정상[OK]')
            self.message_signal.emit(f" 시리얼 {self.proj_port_name} Port 정상[OK]")
        else:
            print(f'시리얼 {self.proj_port_name} Port 에러[ERR]')
            self.message_signal.emit(f" 시리얼 {self.proj_port_name} Port 에러[ERR]")
        print('*******************************')

        # 시리얼 검사 로그 기록
        log_text = QDateTime.currentDateTime().toString("\nyyyy-MM-dd HH:mm.ss |")     # 로그. 시작 시간 기록
        if ret_serial_init:
            log_text = log_text + ' Serial[START TEST OK]'
        else:
            log_text = log_text + ' Serial[START TEST ERR]'
        self.sch_Log_Save(log_text)
        
        # [일반 스케줄 타이머 시작]
        self.sch_Timer_Start()

        while True:
            time.sleep(SCH_SLEEP_SEC)
            # print(".")

            if TEST or KEY_TEST:
                print(".")

            # 키보드 제어 테스트 모드
            #  멈추어 있는다. 타이머 동작을 위해서 계속 다른 키를 입력해야 한다.
            if KEY_TEST:
                event = keyboard.read_event()
                if event.event_type == keyboard.KEY_DOWN:
                    key = event.name.lower()    # 소문자로 변환
                    if key == 'o':
                        print('ON')
                        self.now_On()
                    elif key == 'f':
                        print('OFF')
                        self.now_Off()
                    elif key == 'd':
                        print('DOWN')
                        self.now_Down()
                    elif key == 'c':
                        print('DOWN candel')
                        self.now_DownCancel()
                    elif key == 'p':
                        if self.sch_quiz_progress_set == True:
                            print('quiz progress SET : False')
                            self.sch_quiz_progress_set = False
                        else:
                            print('quiz progress SET : True')
                            self.sch_quiz_progress_set = True

            self.sleep_Timeout()
        ###########################################################################


    ############################################################
    # 스케줄 설정 데이터 가져오기
    def schedule_Int_Load(self):

        # 현재 날짜와 시간 가져오기
        now = datetime.datetime.now()

        # 요일을 정수로 얻기 (월요일: 0, 화요일: 1, ... 일요일: 6)
        weekday_number = now.weekday()

        # 정수 요일을 문자열로 변환
        weekday_name = WEEKDAY_NAMES[weekday_number]
        print(f' 현재 요일은 {weekday_name}')

        # 1. 엑셀파일 읽기
        self.read_wb = load_workbook(f'{SCHEDULER_PATH}{SCHEDULER_FILE}', data_only=True)

        self.read_ws = self.read_wb.sheetnames     # 리스트 형태로 반환
        self.read_ws = self.read_wb[self.read_ws[0]]
        print(self.read_ws)


        # 2.정보 가져오기
        self.proj_port_name = self.read_ws.cell(row=PROJ_COM_NAME[0], column=PROJ_COM_NAME[1]).value
        self.proj_port_speed = self.read_ws.cell(row=PROJ_COM_SPEED[0], column=PROJ_COM_SPEED[1]).value
        print(self.proj_port_name, self.proj_port_speed)

        #   프로젝터 명령어
        self.cmd = {}
        self.cmd['pwr_on'] =  self.read_ws.cell(row=PROJ_PWR_ON_SEND_CMD[0], column=PROJ_PWR_ON_SEND_CMD[1]).value
        self.cmd['pwr_on_rec'] =  self.read_ws.cell(row=PROJ_PWR_ON_REC_CMD[0], column=PROJ_PWR_ON_REC_CMD[1]).value
        self.cmd['pwr_off'] =  self.read_ws.cell(row=PROJ_PWR_OFF_SEND_CMD[0], column=PROJ_PWR_OFF_SEND_CMD[1]).value
        self.cmd['pwr_off_rec'] =  self.read_ws.cell(row=PROJ_PWR_OFF_REC_CMD[0], column=PROJ_PWR_OFF_REC_CMD[1]).value
        # print('프로젝터 명령어 : ', self.cmd)

        #   time
        # self.on_pre_time = self.read_ws.cell(row=SCH_PWR_ON_PRIE_TIME[0], column=SCH_PWR_ON_PRIE_TIME[1]).value
        # self.off_after_time = self.read_ws.cell(row=SCH_PWR_OFF_AFTER_TIME[0], column=SCH_PWR_OFF_AFTER_TIME[1]).value
        # self.pc_off_delay = self.read_ws.cell(row=SCH_SHUTDOWN_DELAY[0], column=SCH_SHUTDOWN_DELAY[1]).value

        #   scheduler table
        #   TEST : 1 Line
        # self.sch_table =  self.read_ws.cell(row=SCH_TIME_TABLE[0], column=SCH_TIME_TABLE[1]).value
        # print(self.on_pre_time, self.off_after_time, self.pc_off_delay, self.sch_table)


        # 1. 미리켜기 시간 가져오기
        self.pre_set_time = self.getTime(SCH_PWR_ON_PRIE_TIME[0], SCH_PWR_ON_PRIE_TIME[1])
        print('pre_time :', self.pre_set_time)

        # 2. 후끄기 시간 가져오기
        self.delay_set_time = self.getTime(SCH_PWR_OFF_AFTER_TIME[0], SCH_PWR_OFF_AFTER_TIME[1])
        print('delay_time :', self.delay_set_time)

        # 3. 셧다운(pc종료) 딜레이 시간
        self.shutdown_delay_set_time = self.getTime(SCH_SHUTDOWN_DELAY[0], SCH_SHUTDOWN_DELAY[1])
        print(self.shutdown_delay_set_time)
        # SET
        self.shutdown_delay_set_qtime = self.toQTime(self.shutdown_delay_set_time)
        self.shutdown_delay_set_sec = self.shutdown_delay_set_qtime.hour() * 3600 + self.shutdown_delay_set_qtime.minute() * 60 + self.shutdown_delay_set_qtime.second()
        print('shutdown_delay_qtime :', self.shutdown_delay_set_qtime)
        # print('shutdown_delay_set_sec :', self.shutdown_delay_set_sec)

        # a.현재시간
        # b.시작 시간
        # c.지연 시간

        # [테스트] 3. 월요일 가져오기 
        # weekday_number = WEEKDAY_NAMES.index('월요일')
        # print(weekday_number, WEEKDAY_CELL[weekday_number], SCH_TIME_TABLE[1] + WEEKDAY_CELL[weekday_number])

        # [스케줄 데이터 가져오기 ]self.daySchedule_data
        self.daySchedule_data = self.getSchDayWeek(SCH_TIME_TABLE[0], SCH_TIME_TABLE[1]+ WEEKDAY_CELL[weekday_number])
        
        print('self.daySchedule_data = ', self.daySchedule_data)

    ############################################################
    # 시리얼 포트 초기화
    def serial_Init(self):
        # self.proj_port_name = 'COM1'
        # print(' self.proj_port_name :', type(self.proj_port_name), self.proj_port_name)
        # print(' self.proj_port_speed :', type(self.proj_port_speed), self.proj_port_speed)
        # print(f' {self.proj_port_name} / {self.proj_port_speed}')

        # 시리얼포트 선택 : 4800, 9600, 19200, 38400, 115200
        if self.proj_port_speed == 4800:
            self.baudrate = QSerialPort.Baud4800
        elif  self.proj_port_speed == 9600:
            self.baudrate = QSerialPort.Baud9600
        elif  self.proj_port_speed == 19200:
            self.baudrate = QSerialPort.Baud19200
        elif  self.proj_port_speed == 38400:
            self.baudrate = QSerialPort.Baud38400
        elif  self.proj_port_speed == 115200:
            self.baudrate = QSerialPort.Baud115200

        # 시리얼 포트 환경변수 설정
        self.port = QSerialPort()
        self.port.setBaudRate( self.baudrate )
        self.port.setPortName( self.proj_port_name )
        self.port.setDataBits( QSerialPort.Data8 )
        self.port.setParity( QSerialPort.NoParity )
        self.port.setStopBits( QSerialPort.OneStop )
        self.port.setFlowControl( QSerialPort.NoFlowControl )

        # 시리얼 포트로 데이터를 수신 가능하게 되었을 때에 발행> readyRead 시그널
        #  https://doc.qt.io/qtforpython/PySide6/QtCore/QIODevice.html
        # self.port.readyRead.connect(self.read_from_port)

        # 시리얼 포트 OPEN
        r = self.port.open(QIODevice.ReadWrite)
        # self.port.write(b'\x04\x03\x02\x01')    #[test]
        # self.port.waitForBytesWritten(1000)

        if not r:
            print(f"\n{self.proj_port_name} Port Open [ERR] : {self.proj_port_speed} ")
            # self.serialRecMsg_signal.emit("Port Open Error")
            # self.serialRecMsg_signal.emit(f"시리얼 '{self.port_number}'포트 연결 오류")
            self.serial_Stop()
            return False
        else:
            print(f"\n{self.proj_port_name} Port Open [OK]  : {self.proj_port_speed} ")
            return True
    
    def serial_Stop(self):
        self.port.close()
        # time.sleep(0.5)
        print(f"{self.proj_port_name} Port Close")

    ############################################################
    # READ | read_from_port()
    def read_from_port(self):
        # self.port.waitForBytesWritten()
        data = self.port.readAll()
        
        # print('read_from_port')
        print(f"serial data recive : {data}")

        # 수신 데이터 self.cmd = {} 에서 확인


    ############################################################
    # WRITE | writeToPort
    ############################################################
    def serial_Write(self, cmd):
        log_text = QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm.ss |")     # 로그. 시작 시간 기록
        rec_ack = None

        # 1. 명령어 데이터 딕션너리 가져오기
        if cmd == 'ON':
            cmd_send = self.cmd['pwr_on']
            cmd_rec = 'pwr_on_rec'
        elif cmd == "OFF":
            cmd_send = self.cmd['pwr_off']
            cmd_rec = 'pwr_off_rec'
        else:
            cmd_send = None
            cmd_rec = None
            print(' 정의되지 않는 시리얼 명령어 입니다.')
        
        # 2. 시리얼 포트 초기화
        ret_serial_init = self.serial_Init()
        if ret_serial_init:

            # 3. 시리얼 데이터 전송
            #  1) [cr] 문자열 찾기
            if '[CR]' in cmd_send:
                _data = cmd_send.split('[CR]')
                # QTimer 와 같이 사용시 발생하는 오류
                #   QObject: Cannot create children for a parent that is in a different thread.
                #   (Parent is QSerialPort(0x22b0cda78b0), parent's thread is ThreadScheduler(0x22b0cdb2290), current thread is QThread(0x22b0cdb2150)
                c = self.port.write(_data[0].encode())
                # self.port.waitForBytesWritten(1000)

                # https://lunikism.com/entry/통신-제어-문자STX-ETX-CR-LF
                cr = '\r'
                c = c + self.port.write(cr.encode())
            else:
                c = self.port.write(cmd_send.encode())

            #  2) 데이터 전송
            self.port.waitForBytesWritten(1000)     # 이게 있어야만, 시리얼 전송이 이루어짐

            print(f'   Serial Write [{cmd}]: {cmd_send}')

            # data = data.encode()  # 문자열을 바이트로 변환하여 전송합니다
            # self.port.write(QByteArray(data))

            # https://doc.qt.io/qtforpython/PySide6/QtSerialPort/QSerialPort.html
            # https://doc.qt.io/qtforpython/PySide6/QtCore/QIODevice.html
            # r = self.port.write(data)                   # 데이터를 시리얼 포트로 전송합니다
            # self.port.waitForBytesWritten(1000)     # 이게 있어야만, 시리얼 전송이 이루어짐

            # print(f' Serial Write : {data}, {r}')

            #  3) 응답 대기
            rec_ack = self.wait_for_ack(cmd_rec)      # ACK 대기
            print(f'   응답 : {rec_ack}')

            # 4. 시리얼 포트 닫기
            self.serial_Stop()
        
        # 3. 배경 음악제어
        if cmd == 'ON' and rec_ack == True:
            self.music_signal.emit('music_allPlay')
        elif cmd == 'OFF' and rec_ack == True:
            self.music_signal.emit('music_stop')


        # 4. 로그 기록
        if ret_serial_init:
            log_text = log_text + f' Serial[OK]  '
        else:
            log_text = log_text + f' Serial[ERR] '

        if cmd == 'ON':
            log_text = log_text + f'| ON   '
        elif cmd == "OFF":
            log_text = log_text + f'| OFF  '
        else:
            log_text = log_text + f'| {cmd} '

            
        if rec_ack == True:
            log_text = log_text + f'| [OK]'
        elif rec_ack == False:
            log_text = log_text + f'| [ERR]'
        else:
            pass

        self.sch_Log_Save(log_text)


        return rec_ack
            


    def wait_for_ack(self, rec_cmd):
        rec_cmd = self.cmd[rec_cmd]
        if '[ASK][CR]' in rec_cmd:
            except_ack1 = bytes.fromhex("06")   # [ACK]16진수 06을 바이트로 변환
            except_ack2 = '\r'.encode()         # [CR]
            except_ack = except_ack1 + except_ack2
        else:
            except_ack = None
        
        # print('except_ack :', except_ack.hex())
        # print(" Waiting for ACK...")

        # https://doc.qt.io/qt-6/qserialport.html#waitForReadyRead
        #  True 반환 : readyRead()신호 발생, 읽을 수 있을 때
        #  False 반환 : 오류발생 작업 시간 초과된 경우
        if self.port.waitForReadyRead(WAIT_FOR_READY_READ): # 1초 동안 ACK 대기
            received_data = self.port.readAll()
            self.port.readLine()

            if except_ack == received_data:
                return True
            else:
                return False
        else:
            return False


    # 시간 더하기
    def addTime(self, _time, _addtime):
        h_time, m_time = _time.split(":")
        hour = int(h_time); minute = int(m_time)
        qtime = QTime(hour, minute)

        ah_time, am_time = _addtime.split(":")
        add_hour = int(ah_time); add_minute = int(am_time)

        qtime = qtime.addSecs(3600*add_hour)
        qtime = qtime.addSecs(60*add_minute)

        # print(time.toString("hh:mm"))
        return qtime.toString("hh:mm"), qtime

    # 시간 빼기
    def subTime(self, _time, _addtime):
        h_time, m_time = _time.split(":")
        hour = int(h_time); minute = int(m_time)
        time = QTime(hour, minute)

        ah_time, am_time = _addtime.split(":")
        add_hour = int(ah_time); add_minute = int(am_time)

        time = time.addSecs(-3600*add_hour)
        time = time.addSecs(-60*add_minute)

        # print(time.toString("hh:mm"))
        return time.toString("hh:mm"), time
    
    # 동작 지연시간 더하기
    # def addDelayTime(self, current_qtime, _item):
    def addDelayTime(self, current_qtime):
        _add_seconds = SCH_QUIZ_PROGRES_DELAY_MIN * 60      # 분을 초단위로 변환
        # new_qtime = _item[Sch.QTIME.value].addSecs(_add_seconds)
        new_qtime = current_qtime.addSecs(_add_seconds)
        return new_qtime.toString('hh:mm'), new_qtime

    # "00:00" -> msec 로 변환
    def toMSecTime(self, _time):
        h_time, m_time = _time.split(":")
        hour = int(h_time); minute = int(m_time)
        time = QTime(hour, minute)
        milliseconds = time.msecsSinceStartOfDay()
        return milliseconds

    # "00:00" -> Qtime type 으로 변환
    def toQTime(self, _time):
        try:
            h_time, m_time = _time.split(":")
            hour = int(h_time); minute = int(m_time)
            time = QTime(hour, minute)
        except:
            time = QTime(0, 0)
        return time

    # 시간 가져오기 - '00:00 형식, 00:00:00 형식(엑셀 시간형식)
    def getTime(self, sch_row, sch_colum):
        # TIME CELL
        cell_time = self.read_ws.cell(row=sch_row ,column=sch_colum).value
        chk_time = re.fullmatch(TIME_PATTERN, str(cell_time))     # 일치하지 않으면 None 반환
        if chk_time != None:
            return cell_time
        elif isinstance(cell_time, datetime.time):  # datetime.time 타입 인지 검사
            # print('datetime.time 타입')
            return cell_time.strftime("%H:%M")      # datetime.time -> str(hh:mm) 문자열로 변환
        else:
            # return None   # 에러 발생
            return 0


    # 초(sec)를 QTime 형식으로 변환
    def secToQtime(self, total_secondes):
        hours = total_secondes // 3600
        minutes = (total_secondes % 3600) // 60
        seconds = total_secondes % 60

        qtime = QTime(hours, minutes, seconds)
        return qtime


    def getSchDayWeek(self, sch_row, sch_colum):
        # return [[ON, '08:30', QTime(8, 30, 0, 0), 0], [, , ,], ... ]

        # 스케쥴 값 가져오기
        list = []
        _index = 0
        for _row in range(sch_row, sch_row+SCH_ROW_MAX):
            err = False
            # ON/OFF CELL
            cell_control = self.read_ws.cell(row=_row ,column=sch_colum).value
            if cell_control != None:
                cell_control = cell_control.upper()
            # TIME CELL
            cell_time = self.read_ws.cell(row=_row ,column=sch_colum+1).value
            
            # 1. 시간 매칭 검사
            #   https://brownbears.tistory.com/506 
            #   _time = re.fullmatch(TIME_PATTERN, cell_time)       # 일치하지 않으면 None 반환
            chk_time = re.fullmatch(TIME_PATTERN, str(cell_time))   # TEXT TIME 타입  정규식과 일치하지 않으면 None 반환
            if chk_time == None:
                if isinstance(cell_time, datetime.time):        # datetime.time 타입 인지 검사
                    cell_time = cell_time.strftime("%H:%M")     # datetime.time -> str(hh:mm) 문자열로 변환
                else:
                    err = True

            # 2. ON,OFF 테이블 검사
            if cell_control not in CONTROL_CMD and cell_control == None:
                err = True

            # 3. err가 없는 경우
            if err == False:
                list.append([cell_control, cell_time])

        # 미리켜기, 지연끄기 시간 연산하기(빼기, 더하기)
        list2 = []

        #################################################################
        for _ctl, _time in list:
            if _ctl == 'ON':
                text_time, q_time = self.subTime(_time, self.pre_set_time)
            elif _ctl == 'OFF':
                text_time, q_time = self.addTime(_time, self.delay_set_time)

            # data set
            SCH_ITEM[Sch.INDEX.value] = _index
            SCH_ITEM[Sch.CTL.value] = _ctl
            SCH_ITEM[Sch.TTIME.value] = text_time
            SCH_ITEM[Sch.QTIME.value] = q_time
            SCH_ITEM[Sch.SEND.value] = 0
            SCH_ITEM[Sch.RESP.value] = 0

            # print(text_time)
            # list2.append([_ctl, text_time, q_time])
            list2.append(SCH_ITEM.copy())
            _index += 1

        # print(list)
        return list2
        #################################################################

  
    def sch_Time_Check(self, current_qtime, sch_q_item):

        # 범위 시작시간, 범위 끝시간
        sch_start_qtime = sch_q_item[Sch.QTIME.value]      # QTime(12, 0, 0, 0) 12h:0m:0s:0ms
        sch_time_step_sec = SCH_TIME_CHECK_STEP_SEC
        sch_end_qtime = sch_q_item[Sch.QTIME.value].addSecs(sch_time_step_sec)

        sch_start_secs = sch_start_qtime.secsTo(current_qtime)
        sch_end_secs = current_qtime.secsTo(sch_end_qtime)

        if sch_start_secs >= 0 and sch_end_secs >= 0:
            # "주어진 시간은 시작 시간과 끝 시간 사이의 시간값입니다.
            # print(' 시간 범위 안')
            return True
        else:
            # 주어진 시간은 시작 시간과 끝 시간 사이의 시간값이 아닙니다.
            # print(' 시간 범위 밖')
            return False
    
    #####################################################################
    def sleep_Timeout(self):
        current_qtime = QTime.currentTime()
        current_qdate = QDateTime.currentDateTime().date()

        # 날짜가 넘거간 경우 초기화
        if self.sch_timer_qdate != current_qdate:
            # 스케쥴 새로 로딩
            self.schedule_Int_Load()
            self.sch_Timer_Start()
            # 날짜가 변경된 경우 -> main 으로 신호 보내기
            # self.daychange_signal.emit(True)    

        # [test]
        if TEST : 
            current_qtime = TEST_CURRENT_TIME

        # [1] 스케쥴 타이머
        if self.sch_timer_set:

            # ### 5초 Time Check ### #
            sch_elapsed_time_sec = self.sch_start_qtime.secsTo(current_qtime)
            # 5초(SCH_SERIAL_SEND_SEC) 넘는 경우 실행
            # print('SCH_SERIAL_SEND_SEC :', SCH_SERIAL_SEND_SEC)
            if sch_elapsed_time_sec >=  SCH_SERIAL_SEND_SEC:
                self.sch_start_qtime =  current_qtime
                if TEST: 
                    self.sch_start_qtime  = TEST_CURRENT_TIME.addSecs(-SCH_SERIAL_SEND_SEC)
            # ###################### #
                # 1.남은 스케줄 검사 [에러, 문제발생]
                self.sch_remaining_list = self.sch_Remaining(current_qtime)    # self.sch_remaining_list = []  남은 스케줄 저장
                sch_all_cnt = len(self.daySchedule_data)
                sch_remaining_cnt = len(self.sch_remaining_list)
                # print(f'전체:{sch_all_cnt}개/남음:{sch_remaining_cnt}개')
                
                # print(f'전체 {sch_all_cnt}개/남은 {sch_remaining_cnt}개 {self.sch_remaining_list}')
                if TEST or KEY_TEST:
                    print(f'전체:{sch_all_cnt}개/남음:{sch_remaining_cnt}개')

                #   스케쥴이 없는 경우
                # print('len(self.sch_remaining_list) :', len(self.sch_remaining_list))
                if len(self.sch_remaining_list) == 0:
                    # [ 종료 시간 설정이 없는 경우] 종료 타이머 사용하지 않음.
                    if self.shutdown_delay_set_sec == 0:
                        return
                    
                    # [ 동작 조건 ]
                    # if self.sch_prev != 'ON':   # 'None'(오늘 스케쥴 없음) or 'OFF'(이전 OFF)
                    if self.sch_prev == None or self.sch_prev == 'OFF':
                        self.sch_Shutdown_Timer_Start()

                # 2. 현재 스케쥴 검사
                else:
                    # 전송후 시간 검사
                    for index, sch_item in enumerate(self.daySchedule_data):
                        # 시간 검사

                        # #### 시간 검사 #### #
                        #                         스케쥴시간 == 현재 시간  and  시리얼 응답 <= 0
                        if self.sch_Time_Check(current_qtime, sch_item):
                            
                            # 전송후 응답 기록이 있으면, 리턴
                            if sch_item[Sch.RESP.value] > 0:
                                return
                            # SERIAL_SEND_MAX 최대 송신 횟수 초과시 리턴
                            if sch_item[Sch.SEND.value] >= SERIAL_SEND_MAX:
                                self.message_signal.emit(f"Serial Write [{self.sch_prev}], 최대 전송 횟수 초과")
                                return
                            
                            # print(f'전체 {sch_all_cnt}개/남은 {sch_remaining_cnt}개 : {self.sch_remaining_list}')
                            
                            # **** (1) 스케쥴 명령어 동작 *****
                            if sch_item[Sch.CTL.value] == 'ON':
                                rec_ack = self.serial_Write('ON')
                                self.sch_prev = 'ON'

                            elif sch_item[Sch.CTL.value] == 'OFF':
                                # 퀴즈 진행중이면
                                if self.sch_quiz_progress_set:
                                    # 시간 지연
                                    _time, _qtime = self.addDelayTime(current_qtime)
                                    sch_item[Sch.TTIME.value] = _time
                                    sch_item[Sch.QTIME.value] = _qtime
                                    time_s = _qtime.toString('hh:mm ss초')

                                    print(f' [스케쥴] 퀴즈 진행중으로 "프로젝터 OFF" 동작을 {SCH_QUIZ_PROGRES_DELAY_MIN}분 후인 {time_s} 에 진행합니다.')
                                    return
                                else:
                                    rec_ack = self.serial_Write('OFF')
                                    self.sch_prev = 'OFF'
                            
                            elif sch_item[Sch.CTL.value] == 'DOWN':
                                # 퀴즈 진행중이면
                                if self.sch_quiz_progress_set:
                                    # 시간 지연
                                    # _time, _qtime = self.addDelayTime(current_qtime, sch_item)
                                    _time, _qtime = self.addDelayTime(current_qtime)
                                    sch_item[Sch.TTIME.value] = _time
                                    sch_item[Sch.QTIME.value] = _qtime
                                    time_s = _qtime.toString('hh:mm ss초')

                                    print(f' [스케쥴] 퀴즈 진행중으로 "PC OFF 동작"을 {SCH_QUIZ_PROGRES_DELAY_MIN}분 후인 {time_s} 에 진행합니다.')
                                    return
                                else:
                                    self.sch_Shutdown_Now()    # 즉시 종료

                            
                            self.daySchedule_data[index][Sch.SEND.value] += 1       # 전송횟수+1
                            print(f'   전송시도 : {self.daySchedule_data[index][Sch.SEND.value]}')

                            self.message_signal.emit(f"프로젝터[{self.sch_prev}], 전송:{self.daySchedule_data[index][Sch.SEND.value]}, 응답:{rec_ack}")

                            # (2) 응답 유무 [리스트]에 기록
                            if rec_ack:
                                self.daySchedule_data[index][Sch.RESP.value] += 1   # 수신횟수+1
                            else:
                                pass


        # [2] PC 종료 타이머 : 매초 검사 - OFF 후 스케쥴이 없거나, 당일 이후 스케쥴이 없는 경우
        if self.sch_shutdown_timer_set:
            # 퀴즈 진행중이면
            if self.sch_quiz_progress_set:
                # (1) shutdown tiemr reset
                self.sch_shutdown_timer_start_qtime = current_qtime     # 종료 시간 reset
                print(f' [타이머] 퀴즈 진행중 PC DOWN 동작 {int(self.shutdown_delay_set_sec/60)}분 후 진행.')
            else:
                # (2) shutdown timer countdown
                #    PC 종료 시간 검사 (예: 5분)
                elapsed_time_sec = self.sch_shutdown_timer_start_qtime.secsTo(current_qtime)      # 지난시간
                # print('elapsed_time_sec :', elapsed_time_sec)
                
                # 초과시 종료 동작
                if elapsed_time_sec >= self.shutdown_delay_set_sec :
                    self.sch_Shutdown_Now()    # 즉시 종료

                # 남은 시간 표시
                else:
                    remaining_qtime = self.shutdown_delay_set_qtime.addSecs(-elapsed_time_sec)
                    print(' PC종료 남은 시간 : ', remaining_qtime.toString('hh:mm ss초'))
                    # 메시지 전송
                    self.message_signal.emit(f"PC종료 남은 시간:{remaining_qtime.toString('hh:mm ss초')}")


    def sch_Timer_Start(self):
        self.sch_timer_qdate = QDateTime.currentDateTime().date()
        # datetime2 = QDateTime(QDate(2023, 9, 10), QTime(21, 0, 23)).date()

        self.sch_shutdown_timer_set = False  # PC종료 타이머 [중지]
        self.sch_timer_set = True            # 스케쥴 타이머 [시작]
        self.sch_start_qtime = QTime.currentTime()
        if TEST: 
            self.sch_start_qtime  = TEST_CURRENT_TIME.addSecs(-SCH_SERIAL_SEND_SEC)
        print(' *일반 타이머 시작')


    def sch_Shutdown_Timer_Start(self):
        self.sch_timer_qdate = QDateTime.currentDateTime().date()

        self.sch_shutdown_timer_set = True   # PC종료 타이머 [시작]
        self.sch_timer_set = False           # 스케쥴 타이머 [중지]
        self.sch_shutdown_timer_start_qtime = QTime.currentTime()
        if TEST: 
            self.sch_shutdown_timer_start_qtime  = TEST_CURRENT_TIME.addSecs(-SCH_SERIAL_SEND_SEC)
            print(' self.sch_shutdown_timer_start_qtime :', self.sch_shutdown_timer_start_qtime)
        print(' *PC종료 타이머 시작')


    def sch_Shutdown_Now(self):
        # 로그 기록
        log_text = QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm.ss | DOWN")     # 로그. 시작 시간 기록
        self.sch_Log_Save(log_text)
        time.sleep(1)   # 파일을 저장하기를 기다림.

        # pc 종료 명령어
        self.sch_shutdown_timer_set = False  # PC종료 타이머 [중지]
        self.sch_timer_set = False           # 스케쥴 타이머 [중지]
        print( " [PC 종료] 'shutdown -s -f'")
        if SHUTDOWN:
            print('** [PC 종료] 시작 **')
            os.system('shutdown -s -f')
            sys.exit()

        self.sch_Shutdown_Timer_Start()


    def sch_Remaining(self, current_qtime):
        # [ 문제점 해결 ] 복사해서 사용하지 않으면,
        #   리스트와 유사한 결과: self.sch_start_qtime = 값이 변경되어 버림
        copy_qtime = QTime(current_qtime)
        copy_qtime.setHMS(copy_qtime.hour(), copy_qtime.minute(), 0 ,0)    # 초, 분 = 0, 0 

        sch_remaining_list = []     # 남은 스케줄 저장

        # all_cnt = len(self.daySchedule_data)
        
        if len(self.daySchedule_data):  # 스케줄이 있는 경우: not 0
            for sch_item in self.daySchedule_data:
                # 남은 스케줄 검색
                # [ 문제점 ]
                # print(f' {sch_item[Sch.QTIME.value]} >= {current_qtime}')
                # PySide6.QtCore.QTime(22, 41, 2, 0) >= PySide6.QtCore.QTime(22, 41, 21, 555)
                if sch_item[Sch.QTIME.value] >= copy_qtime:
                    # print(sch_item[Sch.TTIME.value])
                    sch_remaining_list.append(sch_item)
            
        #    남은 스케줄 카운트, 남은 스케줄 리스트
        # remaining_cnt = len(self.sch_remaining_list)
        return sch_remaining_list
    

    def now_On(self):
        rec_ack = self.serial_Write('ON')
        self.sch_prev = 'ON'
        print(f' 수동 ON 응답 : {rec_ack}')
        self.message_signal.emit(f"단축키: 프로젝터 [{self.sch_prev}], 응답:{rec_ack}")
        self.sch_Timer_Start()

    def now_Off(self):
        rec_ack = self.serial_Write('OFF')
        self.sch_prev = 'OFF'
        print(f' 수동 OFF 응답 : {rec_ack}')
        self.message_signal.emit(f"단축키: 프로젝터 [{self.sch_prev}], 응답:{rec_ack}")

    def now_Down(self):
        print(f' 수동 PC 종료 명령 타이머 실행')
        self.sch_Shutdown_Timer_Start()
        self.message_signal.emit(f"단축키: PC 종료 타이머 시작")

    def now_DownCancel(self):
        print('PC 종료동작을 타이머 중지')
        self.sch_Timer_Start()
        self.message_signal.emit(f"단축키: PC 종료 타이머 중지")

    def now_Down_Togle(self):
        if self.sch_shutdown_timer_set:
            self.now_DownCancel()
            
            # [문제점] 스케쥴이 없는 경우: 종료 타이머 계속 자동 시작하는 문제가 있음
            if len(self.sch_remaining_list) == 0:
                self.sch_prev = 'ON'   # [문제 해결] 종료 타이머 동작 못하도록 한다.
        else:
            self.now_Down()

    def sch_Log_Save(self, log_text):
        log_text = log_text + '\n'
        # 현재 날짜와 시간 가져오기
        current_datetime = datetime.datetime.now()
        # 월별 파일명 생성 (예: 2023-09.txt)
        file_name = current_datetime.strftime(f"{SCH_LOG_FILE_NAME}%Y-%m.txt")
        # 파일 경로 지정
        file_path_name = f'{LOG_PATH}{file_name}'
        # 파일에 텍스트 작성
        with open(file_path_name, "a") as file:
            file.write(log_text)


    #####################################################################



class MainApp:
    def __init__(self):
        # Qtimer, QSerialPort 는 쓰레드 내에서 실행해야 한다.
        #   PyQt UI 사용시 쓰레드 생성되어 동작하는 듯함.

        #  1) create the video capture thread
        self.threadScheduler = ThreadScheduler()

        #  2) signal connect
        self.threadScheduler.message_signal.connect(self.message_slot)
        
        #  3) start the thread
        self.threadScheduler.start()

        # threadScheduler.schedule_Int_Load()


    @Slot(str)
    def message_slot(self, msg):
        print(f'   message: {msg}')



if __name__ == "__main__":
    import keyboard     # 테스트용

    app = QCoreApplication([])
    main_app = MainApp()
    sys.exit(app.exec())

    # ThreadScheduler = ThreadScheduler()
    # ThreadScheduler.run()


'''

1. 문제 검증
   S: 선택형  - 문항번호, 문제, 답안 2개 이상 있는지 검사
   O: OX 형   - 문항번호, 문제, 답안 2개 ( o, x 중 1개만 있어야 함.)
    -> 선택형 문항 갯수
    -> ox형 문항 갯수
    -> 영역별 문항 갯수

2. 정답 순서 바꾸기 
   S: 선택형만 - 답안을 리스트화 답1:   [ ['A', '안녕], ['','바보' ], ['', '애정'], ['', '그림'] ] 
              -> 리스트 랜덤화       [ ['','바보' ], ['A', '안녕], ['', '그림'], ['', '애정'] ]
              -> A 위치 찾기 -> 답2
   리스트 파일로 저장

3. 문제 순서 바꾸기
   [ [1, '문제', '답안1', 답안2, 답안3], [2, '문제', '답안1', 답안2, 답안3]]
     -> 리스트 랜덤화
   리스트 파일로 저장

 과목별 -> 문제 갯수
 과목선택 -> 과목벌 저장? 모두 저장? 
     gen_quiz_all_dic = {'종합','qt.dat'}
     gen_quiz_part_dic = {'유퀴즈': 'q1.dat', '수학': 'q2.dat', '영어':'q3.dat'}

'''


from openpyxl import Workbook, load_workbook
from setup import *
import random, os
import val
import fnmatch
from openpyxl_image_loader import SheetImageLoader
from openpyxl.drawing.image import Image as OImage
from PIL import Image as PImage
from openpyxl.utils import get_column_letter
import copy

import sqlite3


DEBUG = False
# DEBUG = True

XLS_START_ROW = 6       # 퀴즈문제 데이터 시작줄
# Q_MAX_ROW = 
XLS_ROW = 0
XLS_NO = 0            # 문항 번호
XLS_TYPE = 1          # 문제타입 OX, C, M
XLS_SUBJECT = 2           # 과목(영역)
XLS_QUESTION = 3          # 문제
XLS_QUESTION_IMAGE = 4    # 문제 이미지
XLS_ANSWER = 5            # 정답
XLS_CHOICE = 6            # 선택1~
XLS_CHOICE_MAX = 100      # 선택형 최대 , 그러나 답은 5번 안에 있어야 함.

XLS_TYPE_CHOICE = 'C'        # 5지 선택형
XLS_TYPE_OX = 'O'            # OX 선택형
XLS_TYPE_MIX = 'M'           # 5지 선택형 생성형
XLS_TYPE_MIX_CNT = 3         # MIX 문항을 만드는데 필요한 최소 문항 갯수

DB_QID = 0      # QuestionID INTEGER PRIMARY KEY,
DB_CID = 1      # CategoryID TEXT NOT NULL,
DB_QMOD = 2     # QuizMode TEXT,
DB_QEXP = 3     # QuestionExpand TEXT,
DB_QNUM = 4     # QuizNumber INTEGER,
DB_QTYPE = 5    # QuizType TEXT,
DB_QAREA = 6    # QuizArea TEXT,
DB_QDEF = 7     # DifficultyLevel INTEGER,
DB_QQ = 8       # QuestionText TEXT,
DB_QA = 9       # CorrectAnswer TEXT,
DB_DATE = 10    # CreteDate TEXT


class XlsxQuizPreLoad():
    # _1. 파일이름 리스트에 저장
    def __init__(self, quiz_select_mode):
    # def xlsxQuizListLoad(self):
        files = os.listdir(QUIZ_PATH)
        # quizFileList = [i for i in files if i.endswith('.xlsx')]

        # 파일 이름 리스트
        if quiz_select_mode == 'day_seq' or 'day_random' or 'seq' or 'random':
            # 파일 이름 패턴 매칭
            #   https://blog.naver.com/PostView.nhn?blogId=hankrah&logNo=221829206710
            quizFileList = [i for i in files if fnmatch.fnmatch(i, 'quiz*.xlsx')]
        elif quiz_select_mode == 'grade' :
            quizFileList = [i for i in files if fnmatch.fnmatch(i, 'gradequiz*.xlsx')]

        # _1. 파일이름 리스트에 저장
        val.quizFileList = quizFileList
        print(f'val.quizFileList = {val.quizFileList}')

        # NEW _2. 퀴즈 제목 가져오기
        val.quizNameList = self.getQuizNameList()
        print(f'val.quizNameList = {val.quizNameList}')


    # _2. fileNum에 해당하는 퀴즈 영역이름 가져오기
    def getQuizInfo(self, fileNum):
        print(f'* {fileNum}')
        fileName = f'{QUIZ_PATH}{val.quizFileList[fileNum]}'

        read_wb = load_workbook(fileName, data_only=True)
        read_ws = read_wb.sheetnames              # 리스트 형태로 반환
        read_ws = read_wb[read_ws[0]]
        
        # 퀴즈 기본 정보 가져오기
        q_area = read_ws.cell(row=Q_AREA_NAME[0], column=Q_AREA_NAME[1]).value    # 영역 이름 
        speedQuizTime = str(read_ws.cell(row=Q_TIME[0], column=Q_TIME[1]).value )     # 제한 시간
        quizTitleImage = read_ws.cell(row=Q_TITLE_IMAGE[0], column=Q_TITLE_IMAGE[1]).value    # 퀴즈 이미지
        quizMode = read_ws.cell(row=Q_QUIZ_MODE[0], column=Q_QUIZ_MODE[1]).value    # 퀴즈 이미지
        quizOption1 = read_ws.cell(row=Q_QUIZ_OPTION1[0], column=Q_QUIZ_OPTION1[1]).value    # 퀴즈 옵션1
        quizOption2 = read_ws.cell(row=Q_QUIZ_OPTION2[0], column=Q_QUIZ_OPTION2[1]).value    # 퀴즈 옵션2
        quizModeWordAfter = read_ws.cell(row=Q_QUIZ_MODE_WORD_AFTER[0], column=Q_QUIZ_MODE_WORD_AFTER[1]).value   # "word quiz = 영단어 퀴즈" 시 질문 문장

        read_wb.close()

        # qfile=QFileData(fileName, self.q_area, self.speedQuizTime, self.quizMode, self.quizModeWordAfter)
        # qFiledb = QuizDAO(qfile)

        print(' getQuizInfo : ', q_area, speedQuizTime, quizTitleImage, quizMode, quizOption1, quizOption2 )
        return q_area, speedQuizTime, quizTitleImage, quizMode, quizOption1, quizOption2
    
    
    # 파일 이름 가져오기
    def getQuizNameList(self):
        quizNameList=[]
        for qfile in val.quizFileList:
            read_wb = load_workbook(f'{QUIZ_PATH}{qfile}', data_only=True)
            read_ws = read_wb.sheetnames     # 리스트 형태로 반환
            read_ws = read_wb[read_ws[0]]
            q_area = read_ws.cell(row=Q_AREA_NAME[0], column=Q_AREA_NAME[1]).value    # 영역 이름 
            read_wb.close()
            quizNameList.append(q_area)

        # 퀴즈 파일 이름 리스트를 파일로 저장 https://noanomal.tistory.com/18
        with open(f'{QUIZ_PATH}{QUIZ_NAME_LIST}', 'w+') as file:
            file.write('\n'.join(quizNameList))  # '\n' 대신 ', '를 사용하면 줄바꿈이 아닌 ', '를 기준으로 문자열 구분함
        
        return quizNameList

   

# class XlsxQuizLoad(QObject):
class XlsxQuizLoad(QThread):
    finished = Signal()

    def __init__(self, fileNum, random_seq_mode):
        super().__init__()
        self.fileNum = fileNum
        self.random_seq_mode = random_seq_mode
        
        ## 문항 검사
        self.qTypeChoice = 0
        self.qTypeOx = 0
        self.row_xls_quizs = []
        self.q_data_err_list = []
        self.q_area = []
        print('XlsxQuizLoad __init__ ', self.fileNum, self.random_seq_mode)

    def run(self):
        # run 시 인수를 가져올 수 없다
        
        # 파일 전처리
        print('START : XlsxQuizLoad =', self.fileNum, self.random_seq_mode)
        self.quizPreSave(self.fileNum, self.random_seq_mode)
        print('END : XlsxQuizLoad')

        self.db_process()

        self.finished.emit()


    def db_process(self):
        ''' ################################################################## '''
        # 데이터 베이스 연결
        self.conn = sqlite3.connect(f'{DB_PATH}{DB_BASE_FILE}') 
        # cursor 생성
        self.cursor = self.conn.cursor()  # db 처음 시작 지점
        # 퀴즈 테이블 생성
        self.create_quiz_db()
        quiz_file = val.quizFileList[val.quizFileNum]
        ############## file DB 작업  ##################
        # DB DAO(Data Access Object) 생성
        qfilevo = QFileData(quiz_file, self.q_area, self.speedQuizTime, self.quizMode, self.quizModeWordAfter)
        qFiledb = QfileDAO(self.conn, quiz_file)
        qFiledb.compare_file(qfilevo)

        #############  quiz DB 작업  #################
        # DB DAO(Data Access Object) 생성
        quizdb = QuizDAO(self.conn, quiz_file)
        quizdb.compare_quiz(self.org_xls_quizs) # 원래 순서대로 db 에 저장
        #############  quiz DB 커어 연결 닫기  #############
        self.cursor.close()
        self.conn.close()
        ''' ################################################################## '''


    # 퀴즈 파일 읽어 오기 및 퀴즈 전처리 / random_seq_mode = 'quiz_random', 'quiz_seq'
    def quizPreSave(self, fileNum, random_seq_mode):
        self.random_seq_mode = random_seq_mode

        ######################################################################
        ## 1. 엑셀파일 읽기
        self.read_wb = load_workbook(f'{QUIZ_PATH}{val.quizFileList[fileNum]}', data_only=True)

        self.read_ws = self.read_wb.sheetnames     # 리스트 형태로 반환
        self.read_ws = self.read_wb[self.read_ws[0]]
        # print(self.read_ws)

        # 퀴즈 기본 정보 가져오기
        self.q_area = self.read_ws.cell(row=Q_AREA_NAME[0], column=Q_AREA_NAME[1]).value    # 영역 이름 
        self.speedQuizTime = str(self.read_ws.cell(row=Q_TIME[0], column=Q_TIME[1]).value )     # 제한 시간
        self.quizTitleImage = self.read_ws.cell(row=Q_TITLE_IMAGE[0], column=Q_TITLE_IMAGE[1]).value    # 퀴즈 이미지
        self.quizMode = self.read_ws.cell(row=Q_QUIZ_MODE[0], column=Q_QUIZ_MODE[1]).value    # 퀴즈 이미지
        self.quizModeWordAfter = self.read_ws.cell(row=Q_QUIZ_MODE_WORD_AFTER[0], column=Q_QUIZ_MODE_WORD_AFTER[1]).value   # "word quiz = 영단어 퀴즈" 시 질문 문장

        ######################################################################
        # I1-1. 이미지가 포함된 문제 찾기 
        # quizNoImgList, self.image_cells = self.getImageList()
        self.image_cells = self.getImageList()
        # if DEBUG: print('이미지가 포함된 문제 quizNoImgList=', quizNoImgList)   #
        if DEBUG: print('이미지가 포함된 문제 self.image_cells=', self.image_cells)
        
        # image_cell_row_dic = { 문제row:인덱스, 문제row:인덱스, ... }
        #       인덱스로 그림의 위치를 찾기 위함.
        idx = 0
        image_cell_row_dic = {} 
        for item in self.image_cells:
            image_cell_row_dic[item[0]] = idx
            idx += 1
        if DEBUG: print('debug6 :', image_cell_row_dic, '/', self.image_cells)



        # I1-2. 이미지 변환에 사용할 임시 파일 리스트 생성
        self.tmpfiles =[]
        for imgFile in self.image_cells:
            file1 = f'{TEMP_PATH}_r{imgFile[ROW_]}c{imgFile[COL_]}_1.png'
            file2 = f'{TEMP_PATH}_r{imgFile[ROW_]}c{imgFile[COL_]}_2.png'
            self.tmpfiles.append([file1, file2])
        if DEBUG: print(self.tmpfiles)

        # 임시 폴더 생성
        if not os.path.isdir(f'{TEMP_PATH}'):
            os.mkdir(f'{TEMP_PATH}')
        

        # I1-3. 이미지 파일 추출 
        #   이미지 리스트 숫자 변경 > 숫자 1 더하기
        #   => getImageFile 를 하기 위함.
        _get_image_cell = []
        for row  in self.image_cells:
            new_row = []
            for item in row:
                new_row.append(item + 1)
            _get_image_cell.append(new_row)
        # if DEBUG: print('_get_image_cell =', _get_image_cell)
        # print(self.image_cells)
        for idx, _file in enumerate(self.tmpfiles):
            self.getImageFile(_get_image_cell[idx], _file[IN_FILE])


        # I1-4. 이미지 파일 사이즈 변경
        for _file in self.tmpfiles:
            self.imageResize(_file[IN_FILE], _file[OUT_FILE])

        print(f"max_row is in row {self.read_ws.max_row} and max_column is in column {self.read_ws.max_column}")
        
        # self.quiz_count = self.read_ws.max_row - XLS_START_ROW + 1
        self.quiz_count = self.read_ws.max_row - XLS_START_ROW

        print('퀴즈 문항 총 개수 : ', self.quiz_count)

        # Mix 용 : 모든 1번 지문의 값을 가져오기
        self.ansRowList = self.qChoiceRowMix_rowList()
        # print('mix 답안 리스트 :', self.ansRowList)

        '''
        #######        xls 에서 퀴즈 파일 불러오기      ####### 
        [주의]리스트 앞부분에 qRow 컬럼 추가 - 랜덤으로 섞는 경우 추척을 위함.
        '''
        # qRow = XLS_START_ROW
        ######################################################################
        for row in self.read_ws.iter_rows(min_row = XLS_START_ROW): # .iter_rows : 데이터를 1개의 열씩 가져와 작업하는 함수
            qRow = row[0].row           # 현재 행의 첫 번째 셀을 기준으로 행 번호 가져오기
            qNo = row[XLS_NO].value
            qType = row[XLS_TYPE].value
            qSubject =  row[XLS_SUBJECT].value
            qQuestion = row[XLS_QUESTION].value
            qQuestionImage = row[XLS_QUESTION_IMAGE].value    # 임시. 수정요망
            qAnswerInt = row[XLS_ANSWER].value
            if DEBUG: print('debug2 :', qNo, qType, qSubject, qQuestion, qAnswerInt)
            
            # 기본 데이터 검사
            # if None == qNo or None == qType or None == qSubject or None == qQuestion or None == qAnswer:
            
            # 응답Mix 타입의 경우 답을 무조건 1로 설정
            if None != qType:
                if qType.upper() == XLS_TYPE_MIX:
                    qAnswerInt = 1

            # 항목 누락 여부 검사
            if None == qNo or None == qType or None == qQuestion or None == qAnswerInt:
                if DEBUG: print('debug3 : None')
                continue


            # choce count
            # https://stackoverflow.com/questions/33541692/how-to-find-the-last-row-in-a-column-using-openpyxl-normal-workbook
            qChoiceList = []
            # XLS_CHOICE_MAX 1~100 개 까지 가져온다.

            for i in range(XLS_CHOICE_MAX):
                try:
                    if row[XLS_CHOICE+i].value != None:
                        qChoiceList.append(row[XLS_CHOICE+i].value)
                    else:
                        # break
                        continue
                except IndexError:
                    # break
                    continue
            

            if DEBUG: print('debug4 :', qChoiceList)
            # 1) 'C' 선택형 검사
            if qType.upper() == XLS_TYPE_CHOICE:
                lst = []
                # 문항번호, 문제, 정답, 선택1, 선택2 -> 모두 만족해야 함.
                try:
                    if (qNo != None) and (qQuestion != None) \
                        and (qAnswerInt != None) and (qChoiceList[0] != None) \
                        and (qChoiceList[1] != None):

                        # 5개의 선택 문항만 가져오기
                        # (1) Answer MIX 하기
                        qAnswerInt, qChoiceMixList = self.qChoiceMix( qAnswerInt, qChoiceList )
                        # (2) 문제 + Answer MIX
                        lst = [qRow, qNo, qType.upper(), qSubject, qQuestion, qQuestionImage, qAnswerInt ] + qChoiceMixList
                        # (3) 문제 리스크에 추가하기
                        self.row_xls_quizs.append(lst)
                        # 카운트
                        self.qTypeChoice += 1
                    else:
                        self.q_data_err_list.append(qNo)
                except:
                    self.q_data_err_list.append(qNo)

                # 에러 리스트 에 문항 번호 저장


            # 2) 'OX' 형 검사
            elif qType.upper() == XLS_TYPE_OX:
                # 문항번호, 문제, 정답 (o,x)-> 모두 만족해야 함.
                if (qNo != None) and (qQuestion != None) \
                    and (qAnswerInt != None) \
                    and(qAnswerInt.upper() == 'O' or qAnswerInt.upper() == 'X' ):
                    
                    lst = [qRow, qNo, qType.upper(), qSubject, qQuestion, qQuestionImage, qAnswerInt.upper()]    # 타입(C,OX), O/X 대문자로 저장
                    self.row_xls_quizs.append(lst)
                    # 카운트
                    self.qTypeOx += 1

                # 에러 리스트 에 문항 번호 저장
                else:
                    self.q_data_err_list.append(qNo)

                
            # 3) 'mix' 형 검사
            elif qType.upper() == XLS_TYPE_MIX:
                if len(self.ansRowList) < XLS_TYPE_MIX_CNT:
                    print(F"문제 ERR 'mix형'문제를 만들 문항이 {XLS_TYPE_MIX_CNT}개 이하로 문항을 만들수 없습니다. ")
                    continue
                lst = []
                # 문항번호, 문제, 정답, 선택1, 선택2 -> 모두 만족해야 함.
                # try:
                if (qNo != None) and (qQuestion != None) \
                    and (qChoiceList[0] != None):

                    # 5개의 선택 문항만 가져오기
                    # qChoiceList = [qChoice1, qChoice2, qChoice3, qChoice4, qChoice5]
                    # (1) Answer MIX 하기
                    qAnswerInt = 1                  # 1번 답 (고정)
                    qChoiceList = [qChoiceList[0]]  # 1번에 정답 넣기(고정)
                    qAnswerInt, qChoiceMixList = self.qChoiceRowMix(qAnswerInt, qChoiceList)
                    # print(f'답: {qAnswer}, 지문 :{qChoiceMixList}')

                    # 'hose'을(를) 한글로 해석하면?
                    if self.quizModeWordAfter != None:
                        qQuestion = f" '{qQuestion}'{self.quizModeWordAfter}"

                    # (2) 문제 + Answer MIX
                    lst = [qRow, qNo, qType.upper(), qSubject, qQuestion, qQuestionImage, qAnswerInt ] + qChoiceMixList
                    # (3) 문제 리스크에 추가하기
                    self.row_xls_quizs.append(lst)
                    # 카운트
                    self.qTypeChoice += 1
                else:
                    self.q_data_err_list.append(qNo)
                # except Exception as e:
                #     print('ERR Mix 검사 :', e)
                #     self.q_data_err_list.append(qNo)

                # 에러 리스트 에 문항 번호 저장
            else:
                self.q_data_err_list.append(qNo)

            # qRow += 1

        self.read_wb.close()

        # 퀴즈 문항수
        val.quiz_total = len(self.row_xls_quizs)

        # 처음 문제 순서
        pre_XLS_NO =[]
        for data in self.row_xls_quizs:
            pre_XLS_NO.append(data[XLS_NO+1])
        # print(f' >처음순서: {pre_XLS_NO}')
            

        '''
        엑셀에서 불러온 문제 저장 변수 : self.XLS_NO
        # self.row_xls_quizs[3] =  [row, 499, 'M','초급 영단어'," 'study group' 단어(숙어)의 뜻은?", None, 2, '밥그릇', '학습 그룹', '미장원', '국가, 나라', '갑자기']
                                     0   0+1  1+1    2+1             3+1                           4+1  5+1   6+1      7+1           8+1      9+1  10+1     11+1
        '''
        self.org_xls_quizs = copy.deepcopy(self.row_xls_quizs)  #  처음 순서 저장을 위함.       

        #################################################################
        ## *** 3. 전체 문제 순서 Mix,  random.shuffle : 리스트 순서 섞기
        if self.random_seq_mode == 'quiz_random':
            random.shuffle(self.row_xls_quizs)
        elif self.random_seq_mode == 'quiz_seq':
            pass

        #################################################################
        ## *** 문제 번호로 어떻게 바뀌었는지 찾는다
        if self.random_seq_mode == 'quiz_random':
            # 바뀐 문제 순서
            post_no = []
            cnt = 0
            for data in self.row_xls_quizs:
                post_no.append(data[XLS_NO])
                cnt += 1
                # print( 'debug9 :', data)

            # print('바뀐 문제 순서 :', post_no)


        #####################################################################
        # 문제를 섞은 후 원래의 리스트 상태로 만들기
        xls_quizs = []  # 엑셀에 저장할 데이터
        qRow_after = []  # 
        for quiz in self.row_xls_quizs:
            qRow_after.append(quiz[0])   # qRow
            xls_quizs.append(quiz[1:12])  # qNo, ...

        self.xls_quizs = xls_quizs
        if DEBUG: print('qRow_after =', qRow_after)

        #######################################################
        ## 4. 문제 파일로 저장하기
        # https://book.coalastudy.com/data_crawling/week4/stage3

        #   워크북(엑셀파일)을 새로 만듭니다.
        write_wb = Workbook()
        #   현재 활성화된 시트를 선택합니다.
        write_ws = write_wb.active
        #   헤더 추가하기
        write_ws.append( [ self.q_area] ) # 영역
        write_ws.append( [ self.speedQuizTime] ) # 시간
        write_ws.append( [ self.quizTitleImage] ) # 이미지
        write_ws.append( ['문항번호', '문제타입', '과목(영역)', '문제', '문제이미지', '정답', '1', '2', '3', '4', '5',])
        
        last_row = write_ws.max_row
        print('last_row = ', last_row)

        ## 퀴즈 저장하기
        for line in self.xls_quizs:
            write_ws.append(line)
        ## 마지막으로 추가된 행의 번호를 얻음

        #######################################################
        # I5. 이미지 삽입하기
        #   image_cell_row_dic = { 문제번호 : 인덱스, } 
        #                                   : 임시파일 인덱스 -> self.tmpfiles[dic.get(key) 
        #   idx ? 고민해 볼것.
        diff = XLS_START_ROW - last_row -1
        print('diff = ', diff)
        print('image_cell_row_dic = ', image_cell_row_dic)
        for key_row in image_cell_row_dic.keys():
            for idx, row in enumerate(qRow_after):
                if key_row == row-1:           # 엑셀파일의 ( 행, 열 )
                    self.insertImage(write_ws, (idx + last_row + diff, XLS_QUESTION_IMAGE+1), self.tmpfiles[image_cell_row_dic.get(key_row)][OUT_FILE])


        #######################################################
        #   워크북(엑셀파일)을 원하는 이름으로 저장합니다.
        print(f'퀴즈 파일 저장 = {TEMP_PATH }{QUIZ_RANDOM_FILE}')
        write_wb.save(f'{TEMP_PATH }{QUIZ_RANDOM_FILE}')
        write_wb.close()


        #######################################################
        # I6. 임시 이미지 삭제하기
        for i in self.tmpfiles:
            for j in i:
                os.remove(j)

        ## 5. 결과 출력
        # print( f' [문제로딩] 선택형:{self.qTypeChoice}, OX형:{self.qTypeOx}, 총문항:{self.qTypeChoice + self.qTypeOx} , 에러문항{self.q_data_err_list}')
        # print( f' [문제로딩] 이미지 포함 문항:{image_cell_row_dic.keys()}')
        # print( f' [문제로딩] quizPreSave END')




    ###########################################################
    # 새로 만들어진 xls 파일에서 > 퀴즈문제 1개씩 로딩 테스트
    def quizRandomLoadTest(self, q_seq):
        q_seq = q_seq + QUIZ_USER_FILE_START_ROW   # 5번째 라인부터 문제 로딩
        
        ## 5. 엑셀파일 읽기
        self.read_wb = load_workbook(f'{QUIZ_RANDOM_FILE}', data_only=True)

        self.read_ws = self.read_wb.sheetnames     # 리스트 형태로 반환
        self.read_ws = self.read_wb[self.read_ws[0]]

        _row = self.read_ws[q_seq]

        test = []
        for cell in _row:
            data = cell.value
            if data != None and type(data) == type('s'):
                # 공백문자 '\xa0' 가 발생하는 겨우 대처
                # https://stackoverflow.com/questions/61442277/unable-to-replace-delete-xa0-from-a-string-in-python-parsing-from-excel
                test.append(data.replace('\xa0', ' '))
            elif data != None:
                test.append(data)

        print(test)


    ############################################################
    # 퀴즈 문제 선택형 섞기
    #   정답은 1~5 사이에 있어야 함.
    def qChoiceMix(self, answerInt, choiceList):
        # 1. [['A', '응답']... ]
        ans = int(answerInt)
        
        temp_lst = []
        for i in range(len(choiceList)):
            if choiceList[i] != None:
                if i+1 == ans:
                    temp_lst.append(['A', choiceList[i]])
                else:
                    temp_lst.append(['', choiceList[i]])

        # 2. 답안 순서 모두 섞기
        random.shuffle(temp_lst)
        # print(temp_lst)

        # 3. 답 위치 찾기
        ret_lst = []
        for j in range(len(temp_lst)):
            if temp_lst[j][0] == 'A':
                ans = j+1
            ret_lst.append(temp_lst[j][1])

        #     답 위치가 5를 넘어 서는 경우, 5 하단으로 이동
        # print('A:', ans, ret_lst)
        if ans > 5:
            old_pos = ans - 1
            new_pos = random.randint(0, 5-1)
            ret_lst[old_pos], ret_lst[new_pos] = ret_lst[new_pos], ret_lst[old_pos]
            ans = new_pos + 1
        # print('B:', ans, ret_lst[:5])
        return ans, ret_lst
    

    ############################################################
    # 퀴즈 문제 선택형 섞기
    #   정답은 1~5 사이에 있어야 함.
    def qChoiceRowMix(self, qAnswerInt, choiceList):
        # 1. [['A', '응답']... ]
        # choiceList[0] <- 고정위치 정답 포함
        pre_lst = self.qChoiceRowMix_AnsRandom(choiceList) # 정답+오답4개
        # print(f'1 답안 선택 List : {temp_lst}')

        temp_lst = []
        # temp_lst = [정답, 오답, 오답, 오답, 오답]
        for i in range(len(pre_lst)):
            if choiceList[i] != None:
                if i+1 == qAnswerInt:
                    temp_lst.append(['A', pre_lst[i]])
                else:
                    temp_lst.append(['', pre_lst[i]])

        # print('2 문제 섞기 전처리 :', temp_lst)

        # 2. 답안 순서 모두 섞기
        random.shuffle(temp_lst)
        # print(temp_lst)

        # 3. 답 위치 찾기
        ret_lst = []
        for j in range(len(temp_lst)):
            if temp_lst[j][0] == 'A':
                ans = j+1
            ret_lst.append(temp_lst[j][1])

        #     답 위치가 5를 넘어 서는 경우, 5 하단으로 이동
        # print('A:', ans, ret_lst)
        if ans > 5:
            old_pos = ans - 1
            new_pos = random.randint(0, 5-1)
            ret_lst[old_pos], ret_lst[new_pos] = ret_lst[new_pos], ret_lst[old_pos]
            ans = new_pos + 1
        # print('B:', ans, ret_lst[:5])
        return ans, ret_lst

    # 모든 1번 지문의 값을 가져와서 리스트화
    def qChoiceRowMix_rowList(self):
        ansRowList = []
        for i in range(XLS_START_ROW, XLS_START_ROW+self.quiz_count):
            cell = self.read_ws.cell(row=i, column=XLS_CHOICE+1).value
            # 값이 있거나, and XLS_TYPE = 'M' 경우 인 경우 문제 라인의 값을 가져온다.
            if cell != None and \
                self.read_ws.cell(row=i, column=XLS_TYPE+1).value.upper() == 'M':
                ansRowList.append(cell)
        # print(f'답 전체 : {ansRowList}')
        return ansRowList

    # 4개의 오답로 런덤으로 추가
    def qChoiceRowMix_AnsRandom(self, choiceList):
        # 문제점 : 문항이 5개 이하가 되면 리턴되지 않는다. self.quiz_count
        if len(self.ansRowList) < 5:    # len(self.ansRowList) = mix 문항의 갯수
            cnt = len(self.ansRowList)
        else:
            cnt = 5
        while len(choiceList) < cnt:
            # 랜덤 오답 가져오기
            index = random.randint(0, len(self.ansRowList)-1)
            value = self.ansRowList[index]

            if value not in choiceList:   # 새로운 수가 중복이 아니면,
                choiceList.append(value)  # 리스트에 추가
        return choiceList

    '''
    ############################################################
    '''

    ############################################################
    # I-2 이미지가 포함된 문제 찾기 
    def getImageList(self):
        # openpyxl 원하는 사진 삭제
        #   https://kiryanchi.github.io/blog/python/openpyxl_image

        # 시트에 들어있는 이미지들을 복사해서 리스트로 만든다
        copy_sheet1_images = self.read_ws._images[:]
        # 어느 셀에 무슨 사진이 있는지 쉽게 알 수 있게 {key: 셀, value: 이미지 정보} 인 딕셔너리를 만듬
        _images = []
        for image in copy_sheet1_images:
            row = image.anchor._from.row 
            col = image.anchor._from.col
            cell = [row, col]
            if col == XLS_QUESTION_IMAGE:
                _images.append(cell)
            # img_data = image._data

        # print(_images)  # 찾은 이미지 출력
        # return _images
        _images.sort(key=lambda x:x[0])

        # imgQuizNo = []
        # for i in _images:
        #     imgQuizNo.append(i[0] -XLS_START_ROW +2)
        # imgQuizNo.sort()

        # xlsx 파일 이미지 로딩 셀 중복 문제가 발생하는 경우(openpyxl_image_loader 모듈 문제)
        #  2차원 리스트 중복 제거 하기
        #  https://inma.tistory.com/132
        # print('_images = ', _images)
        _image_cells = [list(t) for t in set(tuple(item) for item in _images)]

        # 2차원 배열 정렬
        # https://asxpyn.tistory.com/75
        _image_cells.sort(key=lambda x:x[0])
        # print('\n_image_cells = ', _image_cells)

        # return imgQuizNo, _images
        # return imgQuizNo, _image_cells
        return _image_cells
    

    # I-3. 이미지 파일 추출
    def getImageFile(self, row_column, imgFile):
        # ##################################################################
        # getImageFile(xlsFile, 'E7', '_temp1_1.png'):
        # image_loader = SheetImageLoader(self.sheet1)
        image_loader = SheetImageLoader(self.read_ws)

        # 숫자 인덱스를 변환하는 방법
        #   column_letter = get_column_letter(column_index)
        column_letter = get_column_letter(row_column[COL_])
        cell_index = f'{column_letter}{row_column[ROW_]}'
        if DEBUG: print(' debug I: 이미지 위치 =', cell_index)

        pre_image = image_loader.get(cell_index)
        pre_image.save(imgFile)


    # I-4. 이미지 파일 사이즈 변경
    def imageResize(self, inImgFile, outImgFile):
        # ##################################################################
        # 이미지 사이즈 변경
        #   https://minorman.tistory.com/4
        #   https://blog.naver.com/werbertido/221870723761
        #   https://yeko90.tistory.com/entry/파이썬-기초-PIL을-이용하여-이미지-크기를-변경하자

        # chatGPT 질문 결과
        # 질문: Pillow 라이브러리를 사용하여 정해진 사이즈를 넘지 않게 비율을 유지하면서, 리사이즈하는 방법은?
        
        # 이미지 열기
        max_width = Q_IMAGE_W
        max_height = Q_IMAGE_H

        load_image = PImage.open(inImgFile)

        # 현재 이미지의 크기
        current_width, current_height = load_image.size

        # 가로와 세로 비율 계산
        width_ratio = max_width / current_width
        height_ratio = max_height / current_height

        # 이미지 크기가 최대 크기를 넘지 않도록 비율 조정
        resize_ratio = min(width_ratio, height_ratio)
        new_width = int(current_width * resize_ratio)
        new_height = int(current_height * resize_ratio)

        # 이미지 크기가 최대 크기를 넘지 않도록 비율 조정
        resize_ratio = min(width_ratio, height_ratio)
        new_width = int(current_width * resize_ratio)
        new_height = int(current_height * resize_ratio)

        # 이미지 리사이즈
        resized_image = load_image.resize((new_width, new_height), PImage.LANCZOS)
        resized_image.save(outImgFile)

    # 리사이즈된 이미지 저장


    # ##################################################################
    # I-5. 이미지 사이즈 변경 후 xls 넣기
    def insertImage(self, sheet, row_column, imgFile):

        img = OImage(imgFile)   # from openpyxl.drawing.image import Image

        # 셀 사이즈 변경
        #   https://dotsnlines.tistory.com/720
        # 숫자 인덱스를 변환하는 방법
        #   column_letter = get_column_letter(column_index)
        sheet.row_dimensions[row_column[0]].height = XLS_CELL_H

        # ws2.column_dimensions['C'].width = XLS_CELL_W
        column_letter = get_column_letter(row_column[1])
        sheet.column_dimensions[column_letter].width = XLS_CELL_W
        
        per = img.width / img.height
        img.height = 50
        img.width = img.height * per

        # C3 위치에 img.png 파일의 이미지를 삽입
        # ws2.add_image(img, "C3")
        cell_index = f'{column_letter}{row_column[0]}'
        sheet.add_image(img, cell_index)


    # 테이블 생성
    def create_quiz_db(self):
        ###### create "quiz" table ######
        try:
            self.cursor.execute("""
                CREATE TABLE IF NOT EXISTS quiz(
                    quiz_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_name TEXT NOT NULL,
                    quiz_number INTEGER,
                    type TEXT,
                    area TEXT,
                    level INTEGER,
                    question TEXT,
                    answer TEXT,
                    crete_date TEXT,
                    FOREIGN KEY(file_name) REFERENCES quiz_file(file_name)
                )
            """)
            print(f"[DB1] create quiz table OK!")
        except Exception as e:
            print(f"[DB1 create quiz table Err] {e}")

        ###### create "quiz_file" table ######
        try:
            self.cursor.execute("""
                CREATE TABLE IF NOT EXISTS quiz_file(
                    file_name TEXT PRIMARY KEY,
                    title TEXT,
                    time TEXT NOT NULL,
                    mode TEXT,
                    expand TEXT,
                    crete_date TEXT      
                )
            """)
            print(f"[DB2] create quiz_file table OK!")
        except Exception as e:
            print(f"[DB2 create quiz_file table Err] {e}")

        ###### create "answer" table ######
        try:
            # self.cursor.execute("""
            #     CREATE TABLE IF NOT EXISTS answer(
            #         answer_id INTEGER PRIMARY KEY AUTOINCREMENT,
            #         user_id INTEGER NOT NULL,
            #         file_name TEXT NOT NULL,
            #         quiz_number INTEGER NOT NULL,
            #         answer TEXT,
            #         answer_time INTEGER,
            #         how_to TEXT,
            #         quiz_repeat INTEGER,
            #         date_time TEXT,
            #         FOREIGN KEY(quiz_number) REFERENCES quiz(quiz_number),
            #         FOREIGN KEY(file_name) REFERENCES quiz_file(file_name),
            #         FOREIGN KEY(user_id) REFERENCES user(id)
            #     )
            # """)
            self.cursor.execute("""
                CREATE TABLE IF NOT EXISTS answer(
                    answer_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    file_name TEXT NOT NULL,
                    quiz_number INTEGER NOT NULL,
                    answer TEXT,
                    answer_time INTEGER,
                    how_to TEXT,
                    quiz_repeat INTEGER,
                    date_time TEXT
                )
            """)
            print(f"[DB3] create answers table OK!")
        except Exception as e:
            print(f"[DB3 create answers table Err] {e}")

        ###### create "quiz_visit" table ######
        try:
            # self.cursor.execute("""
            #     CREATE TABLE IF NOT EXISTS quiz_visit(
            #         visit_id INTEGER PRIMARY KEY AUTOINCREMENT,
            #         user_id INTEGER NOT NULL,
            #         file_name TEXT NOT NULL,
            #         date_time TEXT,
            #         FOREIGN KEY(user_id) REFERENCES user(id),
            #         FOREIGN KEY(file_name) REFERENCES quiz_file(file_name)
            #     )
            # """)
            self.cursor.execute("""
                CREATE TABLE IF NOT EXISTS quiz_visit(
                    visit_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    file_name TEXT NOT NULL,
                    file_name_quiz_count INTEGER,
                    quiz_count INTEGER,
                    right_count INTEGER,
                    wrong_count INTEGER,
                    pass_count INTEGER,
                    score INTEGER,
                    date_time TEXT
                )
            """)
            print(f"[DB4] create quiz_visit table OK!")
        except Exception as e:
            print(f"[DB4 create quiz_visit table Err] {e}")

#######################################################################################
# 데이터 베이스 연결
#######################################################################################
        

class QFileData:
    def __init__(self, file_name=None, title=None, time=None, mode=None, expand=None, crete_date=None):
        self.file_name = file_name
        self.title = title
        self.time = time
        self.mode = mode
        self.expand = expand
        self.crete_date = crete_date
    def __str__(self):
        return f"QFileData(file_name={self.file_name}, title={self.title}, time={self.time}, mode={self.mode}, expand={self.expand}, crete_date={self.crete_date})"

class QfileDAO:
    def __init__(self, conn, file_name):
        self.conn = conn
        self.cursor = conn.cursor()
        self.file_name = file_name
    
    def create(self, vo):
        CreteDate = QDate.currentDate().toString('yyyy-MM-dd')
        self.cursor.execute("INSERT INTO quiz_file(file_name, title, time, mode, expand, crete_date) VALUES(?, ?, ?, ?, ?, ?)",
                       (self.file_name, vo.title, vo.time, vo.mode, vo.expand, CreteDate) )
        self.conn.commit()
        print(f'quiz_file DB ADD : {self.file_name}')

    def update(self, vo):
        CreteDate = QDate.currentDate().toString('yyyy-MM-dd')
        try:
            self.cursor.execute("UPDATE quiz_file SET title=?, time=?, mode=?, expand=?, crete_date=? WHERE file_name=?",
                        ( vo.title, vo.time, vo.mode, vo.expand, CreteDate, vo.file_name) )
            self.conn.commit()
            # print(f'Quiz with quiz_id={vo.quiz_id} | quiz_number={vo.quiz_number} updated.')
        except Exception as e:
            print(f'FIE DB ERR update :{e}')

    def read_all(self):
        result = []
        self.cursor.execute("SELECT * FROM quiz_file")
        rows = self.cursor.fetchall()
        for row in rows:
            vo = QFileData(*row)
            result.append(vo)
        return result
    
    def delete(self, vo):
        self.cursor.execute("DELETE FROM quiz_file WHERE file_name=?", (vo.file_name,))
        self.conn.commit()
        print(f'DB DEL file = {vo.file_name}')
    
    def read(self, vo):
        try:
            self.cursor.execute("SELECT * FROM quiz_file WHERE file_name=?", (vo.file_name,))
            rows = self.cursor.fetchone()
            _vo = QFileData(*rows)
            return _vo
        except Exception as e:
            print(f'DB ERR read : {e}')
            _vo = QFileData()
            return _vo

    def compare_file(self, vo):

        _xls_fileList = [vo.file_name, vo.title, vo.time, vo.mode, vo.expand]
        _read = self.read(vo)
        _db_fileList = [_read.file_name, _read.title, _read.time, _read.mode, _read.expand]

        print('_xls_fileList= ', _xls_fileList)
        print('_db_fileList=', _db_fileList)

        if _xls_fileList[0] == _db_fileList[0]:
            print(f"file DB 에 '{_xls_fileList[0]}' 이 존재함")
            if _xls_fileList != _db_fileList:   # 변경 사항 있는 경우
                self.update(vo)
        else:
            print(f"file DB 에 '{_xls_fileList[0]}' 를 추가함")
            self.create(vo)

# VO(Value Object) : 값을 표현하는 클래스      
class QuizData:
    def __init__(self, quiz_id=None, file_name=None, quiz_number=None, type=None, area=None, level=None, question=None, answer=None, crete_date=None):
        self.quiz_id = quiz_id          # 0
        self.file_name = file_name      # 1
        self.quiz_number = quiz_number  # 2
        self.type = type                # 3
        self.area = area                # 4
        self.level = level              # 5
        self.question = question        # 6
        self.answer = answer            # 7
        self.crete_date = crete_date
    
    def __str__(self):
        return f"QuizData(quiz_id={self.quiz_id}, file_name={self.file_name}, quiz_number={self.quiz_number}, type={self.type}, area={self.area}, level={self.level}, question={self.question}, answer={self.answer}), self.crete_date={None}"

# DAO(Data Access Object): 데이터베이스에 접근하여 조작하는 기능을 정의하는 클래스 
class QuizDAO:
    def __init__(self, conn, file_name):
        self.conn = conn
        self.cursor = conn.cursor()
        self.file_name = file_name

    def create(self, vo):
        CreteDate = QDate.currentDate().toString('yyyy-MM-dd')
        self.cursor.execute("INSERT INTO quiz( file_name, quiz_number, type, area, level, question, answer, crete_date) VALUES(?, ?, ?, ?, ?, ?, ?, ?)",
                       (self.file_name, vo.quiz_number, vo.type, vo.area, vo.level, vo.question, vo.answer, CreteDate) )
        # self.conn.commit()
        # print("New quiz added.")

    def update(self, vo):
        CreteDate = QDate.currentDate().toString('yyyy-MM-dd')
        try:
            self.cursor.execute("UPDATE quiz SET file_name=?, quiz_number=?, type=?, area=?, level=?, question=?, answer=?, Crete_date=? WHERE quiz_id=?",
                        (self.file_name, vo.quiz_number, vo.type, vo.area, vo.level, vo.question, vo.answer, CreteDate, vo.quiz_id) )
            # self.conn.commit()
            # print(f'Quiz with quiz_id={vo.quiz_id} | quiz_number={vo.quiz_number} updated.')
        except Exception as e:
            print(f'ERR update :{e}')

    def read_all(self):
        result = []
        try:
            self.cursor.execute("SELECT * FROM quiz WHERE file_name=?", (self.file_name,) )
            rows = self.cursor.fetchall()
            # print(rows)
            for row in rows:
                vo = QuizData(*row)
                result.append(vo)
            # print(f"[DB4 read_All] {result}")
            return result
        except Exception as e:
            print(f'QUIZ DB ERR : {e}')
            return result
    
    def delete(self, vo):
        self.cursor.execute("DELETE FROM quiz WHERE quiz_id=?", (vo.quiz_id,))
        # self.conn.commit()
        # print(f"Quiz with quiz_id={vo.quiz_id} | quiz_number={vo.quiz_number} deleted.")

    def commit(self):
        self.conn.commit()


    '''
    엑셀에서 불러온 문제 저장 변수 : self.xls_quizs
    '''
    def compare_quiz(self, xls_quizs):
        # xls 파일의 내용와 현재 db의 내용을 검사하여 처리

        ## 1. db 에 저장된 퀴즈 가져오기 -> _db_quizs 
        pre_db_quizs = self.read_all()

        _db_quizs = []
        _db_numbers = []
        for _quiz in pre_db_quizs:
            # db 에서 "비교에 필요한 필드"만 추출
            _db_quizs.append([_quiz.quiz_id, _quiz.quiz_number, _quiz.type, _quiz.area, _quiz.level, _quiz.question, _quiz.answer])
            _db_numbers.append(_quiz.quiz_number)

        ## 2. xls_quizs 에서 "비교에 필요한 필드"만 추출 -> _xls_quizs
        _xls_quizs = []
        _xls_numbers = []
        for x_quiz in xls_quizs:

            # xls 에서 필요한 필드만 추출
            i = str(x_quiz[5])
            if i.isdigit():
                x_anser = x_quiz[x_quiz[5]+5]   # 답값의 위치 찾기
            else: # 문자열 이라면
                x_anser = x_quiz[5]             # 현재의 값을 답으로
            qnumber=x_quiz[0]; qtype=x_quiz[1]; qarea=x_quiz[2]; qlevel=None; question=x_quiz[3];  anser=x_anser
            x_quiz = [None, qnumber, qtype, qarea, qlevel, question, anser]
            # print('x_quiz=', x_quiz)
            _xls_quizs.append(x_quiz)
            _xls_numbers.append(qnumber)

        # print(f'_db_numbers = {_db_numbers}')
        # print(f'_xls_numbers = {_xls_numbers}')

        def quiz_number_find(number, _numbers):
            if number in _numbers:
                return True
            else:
                return False    
        
        ## xls 내용과 db 내용을 비교하기
        for in_quiz_xls in _xls_quizs:
            # 1. xls 에 있고
            for in_quiz_db in _db_quizs:
                # qnumber 찾기
                compare = None
                if in_quiz_xls[1] == in_quiz_db[1]: # qnumber [퀴즈 번호]를 찾아서 나머지 비교
                    if in_quiz_xls[1:] == in_quiz_db[1:]:
                        compare = True
                    else:
                        compare = False
                # 1-1) [퀴즈 번호] xls 에 있고, db 내용이 다르면, db update
                if compare == False:
                                            # quiz_id,    fine_name , quiz_number ,  type,          ,area          , level        , question       , answer
                    self.update(QuizData(in_quiz_db[0], self.file_name, in_quiz_xls[1], in_quiz_xls[2], in_quiz_xls[3], in_quiz_xls[4], in_quiz_xls[5], in_quiz_xls[6]))
            # 1-2)  xls에 있고, db 에 없으면, db insert
            if quiz_number_find( in_quiz_xls[1], _db_numbers) == False:
                #                        # quiz_id,    fine_name , quiz_number ,  type,          ,area          , level        , question       , answer
                self.create(QuizData(None, self.file_name, in_quiz_xls[1], in_quiz_xls[2], in_quiz_xls[3], in_quiz_xls[4], in_quiz_xls[5], in_quiz_xls[6]))

        for in_quiz_db in _db_quizs:
            # 2) db 에 있고, xls 없으면, db 삭제
            if quiz_number_find( in_quiz_db[1], _xls_numbers) == False:
                self.delete(QuizData(in_quiz_db[0], None, in_quiz_db[1]))

        self.commit()


''' ################################################################## '''


''' ################################################################## '''


if __name__=="__main__":
    DEBUG = False
    # DEBUG = True

    # 1. 인스턴스 생성
    xlsxQuizPreLoad = XlsxQuizPreLoad('seq')
    # 2. xls 문제 정보 로딩
    xlsxQuizPreLoad.getQuizInfo(0) # xls 리스트 첫번째 문제 정보 로딩

    # 3. 문제 전처리 저장
    xlsxQuizLoad = XlsxQuizLoad(0, 'quiz_random')   # quiz_seq, quiz_random
    xlsxQuizLoad.run()

    print(' test 1 :', xlsxQuizPreLoad.getQuizInfo(0))

    # print(xlsxQuizLoad.getQuizInfo(1))
    # xlsxQuizLoad.quizRandomLoadTest(1)


    '''  m_quiz.py 에 들어갈 내용 #################################    '''
    # # 데이터 베이스 연결
    # conn = sqlite3.connect(f'{DB_PATH}{DB_BASE_FILE}') 
    # # cursor 생성
    # cursor = conn.cursor()  # db 처음 시작 지점
    # # 퀴즈 테이블 생성
    # xlsxQuizLoad.create_quiz_db()
    # quiz_file = val.quizFileList[val.quizFileNum]

    # ############## file DB 작업  ##################
    # # DB DAO(Data Access Object) 생성
    # qfilevo = QFileData(quiz_file, xlsxQuizPreLoad.q_area, xlsxQuizPreLoad.speedQuizTime, xlsxQuizPreLoad.quizMode, xlsxQuizPreLoad.quizModeWordAfter)
    # qFiledb = QfileDAO(conn, quiz_file)
    # qFiledb.compare_file(qfilevo)


    # #############  quiz DB 작업  #################
    # # DB DAO(Data Access Object) 생성
    # quizdb = QuizDAO(conn, quiz_file)
    # quizdb.compare_quiz(xlsxQuizLoad.org_xls_quizs) # 원래 순서대로 db 에 저장

    # # 데이터 추가
    # # quizdb.create(QuizData(None, quiz_file, 2, "타입", '영역', None, "질문?", "답은"))

    '''   #################################   '''
    





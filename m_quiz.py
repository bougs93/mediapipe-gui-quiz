# OpenCV(Python) + PyQt
#  https://blog.xcoda.net/104


'''
** How to display opencv video in pyqt apps **
    https://gist.github.com/docPhil99/ca4da12c9d6f29b9cea137b617c7b8b1 (기본)
    https://webnautes.tistory.com/1290 (보조)


    https://webnautes.tistory.com/1290
    https://blog.xcoda.net/104
Pyqt5와 OpenCV 연동시 주의점
    https://hagler.tistory.com/190


23.4.24 - 웹카메라 인식

정답/오답 기호로 사용할 폰트 : 
    MV Boli
    Comic Sans MS <선택> O,X: 300 size,  PASS: 100 size
'''

####### 사용자 퀴즈 데이터 형식 지정 #######
# user_quiz_data =[]
# [0]학번 /사용자명 / 시작 시간 / 풀이시간
# [1]순번     :
# [2]문항번호  :
# [3]답       :
# [4]Y/N/P    :
# [5]응답     :

# SEQ = 1; QNO = 2; ANS = 3; YNP = 4; REQ = 5; 

# self.quiz_line[]
# Q_NO = 0            # 문항 번호
# Q_TYPE = 1          # 문제타입 OX, C
# Q_TYPE_CHOICE = 'C'
# Q_TYPE_OX = 'OX'
# Q_SUBJECT = 2       # 과목(영역)
# Q_QUESTION = 3      # 문제
# Q_ANSWER = 4        # 정답
# Q_CHOICE = 5        # 선택1~
# Q_START_ROW = 3  # 데이터 시작줄


######################################
# 카메라 OpenCV -> QImage 로 뛰우기
######################################
from PySide6 import QtGui
from PySide6.QtWidgets import *
from PySide6.QtGui import *
from PySide6.QtMultimedia import *
import sys

from PySide6.QtCore import *
import mediapipe as mp
import numpy as np
from openpyxl import Workbook, load_workbook
from circular_progress import CircularProgress
from key_press_timer import Key3PressTimer, FingerKeyTimer
from xlsx_quiz_load import *

from openpyxl_image_loader import SheetImageLoader
from openpyxl.utils import get_column_letter
from PIL.ImageQt import ImageQt

import val
from setup import *


from ini_file import *

###############
from thread_video import *

## .ui -> .py ##
from pyside6_uic import PySide6Ui
from ui_quiz import Ui_quizView
from ui_hand_finger import Ui_hand_widget
HAND_WIDGET_SIZE_W = 500
HAND_WIDGET_SIZE_H = 500


# XLS_START_ROW = 6       # 퀴즈문제 데이터 시작줄
# Q_MAX_ROW = 
Q_ROW = 0
Q_NO = 0            # 문항 번호
Q_TYPE = 1          # 문제타입 OX, C, M
Q_SUBJECT = 2           # 과목(영역)
Q_QUESTION = 3          # 문제
Q_QUESTION_IMAGE = 4    # 문제 이미지
Q_ANSWER = 5            # 정답
Q_CHOICE = 6            # 선택1~
Q_CHOICE_MAX = 100      # 선택형 최대 , 그러나 답은 5번 안에 있어야 함.

Q_TYPE_CHOICE = 'C'        # 5지 선택형
Q_TYPE_OX = 'O'            # OX 선택형
Q_TYPE_MIX = 'M'           # 5지 선택형 생성형
Q_TYPE_MIX_CNT = 3         # MIX 문항을 만드는데 필요한 최소 문항 갯수



class HandWidget(QWidget, Ui_hand_widget):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

class HandCountView(QWidget):
    def __init__(self):
        super().__init__()
        self.graphicsview = QGraphicsView()
        self.scene = QGraphicsScene(self.graphicsview)
        self.graphicsview.setScene(self.scene) 

        self.graphicsview.setStyleSheet("background: transparent; ")

        self.graphicsview.setFrameShape(QGraphicsView.NoFrame)
        self.graphicsview.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.graphicsview.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.handWidget = HandWidget()
        self.handWidget.setStyleSheet("background-color: transparent; ")

        self.proxy = QGraphicsProxyWidget()
        self.proxy.setWidget(self.handWidget)
    
        # 중앙을 변환 기준점으로 설정
        self.proxy.setTransformOriginPoint(self.proxy.boundingRect().center())
        # 하단 중앙을 변환 기준점으로 설정
        # self.proxy.setTransformOriginPoint(self.proxy.boundingRect().center().x(), self.proxy.boundingRect().bottom())
        self.scene.addItem(self.proxy)

        # 나머지 설정(1)
        layout = QVBoxLayout(self)
        layout.addWidget(self.graphicsview)

        #################################################
        self.lb_fingerText = self.handWidget.lb_fingerText
        self.pb_fingerText = self.handWidget.pb_fingerText
        self.lb_circle = self.handWidget.lb_circle

        self.lb_fingerText.setText('')
        self.pb_fingerText.setValue(0)
        self.pb_fingerText.setFormat('')

    def setZoom(self, zoom):
        # 줌
        #########################################################################
        # QGraphicsItem 크기 조정 시 변환 지점 설정 -> scale 중앙점 기준 줌 효과
        # https://stackoverflow.com/questions/69436810/setting-transformation-point-on-scaling-qgraphicsitem
        # (문제점) 손이 사라질 위치에서 위치가 변화됨.
        # self.graphicsview.setFixedSize(HAND_WIDGET_SIZE_W, HAND_WIDGET_SIZE_H)
        # origin = self.proxy.transformOriginPoint()
        # self.proxy.resetTransform()     # 중요: 원래 사이즈로 복원(추가)
        # transform = self.proxy.transform().translate(origin.x(), origin.y())
        # transform.scale(zoom, zoom)
        # transform.translate(-origin.x(), -origin.y())
        # self.proxy.setTransform(transform)

        #########################################################################
        # m_quiz_mental.py 에서 사용한 방법 -> m_quiz.py 에서는 스케일 중앙점 변화가 없음.
        self.graphicsview.setFixedSize(HAND_WIDGET_SIZE_W, HAND_WIDGET_SIZE_H)
        self.graphicsview.resetTransform()      # 원래의 사이즈로 복원(이상 동작)
        self.graphicsview.scale(zoom, zoom)     #    1.0 기준
        self.proxy.setTransformOriginPoint(self.proxy.boundingRect().center())


    def setMove(self, _x, _y, radio):
        size = self.handWidget.size()
        if radio < 1:
            _x = int(_x - size.width()/2)
            _y = int(_y - size.height()/2)
        else:
            # 커지는 화면에서 위치가 심하게 변경되는 것을 보정
            _x = int(_x - (size.width()*radio)/2)
            _y = int(_y - (size.height()*radio)/2)

        self.move( _x, _y)




class QuizDisplay(QWidget, Ui_quizView):

    sound_play_signal = Signal(str)
    main_to_signal = Signal(str)

    quiz2_loadReady_finished = Signal()

    def __init__(self):
        super().__init__()
        self.setupUi(self)  # Ui_quizView

        self.setWindowTitle("Qt live label demo")
        self.lb_no.hide()
        self.lb_type.hide()
        self.lb_code.hide()

        self.lb_question.show()
        self.lb_quizResult.hide()   # 결과
        self.lb_chAnswer.hide()     # 1) ~ 5)

        self.lb_waitMsg.hide()

        self.displayInt()

        # 테스트 에러 방지
        self.keyOneInput = True

        ##### 버전 정보 표시 #####
        try:
            self.lb_verInfo.setText( f'프로그램 {PROGRAM_VER}  |  퀴즈 {QUIZ_VER}  |  Program developer : {PROGRAM_DEVELOPER}')
        except:
            pass

        # https://www.geeksforgeeks.org/pyqt5-access-the-size-of-the-label/
        imageView_width, imageView_height = self.lb_imageView.size().width(), self.lb_imageView.size().height()
        print('frameView Size:',imageView_width, imageView_height)

        ##### Create Circular Progress
        self.progress = CircularProgress()
        self.progress.value = 0
        self.progress.suffix = "%"
        self.progress.font_size = 35
        self.progress.text_color = 0xffffff
        self.progress.width = 300
        self.progress.height = 300
        self.progress.progress_width = 20
        self.progress.progress_color = 0xffffff
        self.progress.progress_rounded_cap = True
        self.progress.add_shadow(True)
        self.progress.text = 'QUIT'
    
        # Center CircularProgress
        self.vlayout = QVBoxLayout()
        self.vlayout.addWidget(self.progress)
        self.center_progress.setLayout(self.vlayout)
    
        ## Key Press Timer 생성 ( timer: 2000ms)
        self.progress.hide()
        self.PressKeyTimer = Key3PressTimer(KEY3_TIMER_INTERVAL, KEY3_PUSH_TIME, KEY3_VIEW_DELAY, self.progress)

        # NEW
        self.createHandViewWidget()
        self.handCountView.hide()

        ## Finger Key Timer 생성
        self.FingerKeyTimer = FingerKeyTimer(FINGER_TIMER_INTERVAL, FINGER_PUSH_TIME, FINGER_VIEW_DELAY, self.handCountView.lb_fingerText, self.handCountView.pb_fingerText )

        # 남은 시간을 가지고 타이머 색상 변경하기
        self.timeWarning = 0

        # 퀴즈 파일 정보(val.quizFileList 에 저장)
        self.xlsxQuizPreLoad = XlsxQuizPreLoad(QUIZ_SELECT_MODE)


        #################################################################################
        # 퀴즈 시작전 메시지 파일 로딩 : def quiz3_ReadyCount_Start(self)
        self.msg_quiz_countdown = None
        try:
            with open(f'{CONFIG_PATH}{MSG_FILE_QUIZ_COUNTDOWN}', 'r', encoding='utf-8') as file:
                txt = file.read()
                # self.lb_msg.setText(txt)
                self.msg_quiz_countdown = txt
        except FileNotFoundError:
            pass



        # 입력 모드 해설 이미지 표시
        #   이미지 비율 맞추기  https://jjoh4803.tistory.com/18
        #   이미지 종횡비       https://bskyvision.com/entry/pyside-keepaspectratio
        self.lb_inputModeImg.setScaledContents(False)   # False 상태에서, 스케일 동작함.

        # ## 손가락 모양표시 테스트 
        # # pixmap = QPixmap(f'{IMG_PATH}{IMG_MODE_FINGER_COUNT_FILE}')
        # # pixmap = QPixmap(f'{IMG_PATH}{IMG_MODE_THUMB_UP_DOWN}')
        # pixmap = QPixmap(f'{IMG_PATH}{IMG_MODE_TOUCH}')
        # pixmap2= pixmap.scaled(self.lb_inputModeImg.size().width(), self.lb_inputModeImg.size().height(), Qt.KeepAspectRatio) 
        # # self.lb_inputModeImg.setAlignment(Qt.AlignCenter)
        # self.lb_inputModeImg.setPixmap(pixmap2)
        
        self.viewInputModImg(None)

        self.test_hint = False
        self.lb_test_hint.setStyleSheet('')

        # 터치 연속 입력 제한
        self.key_repeat_ready = True

        # 퀴즈 모드 이미지
        self.lb_mode_img.setScaledContents(False)   # False 상태에서, 스케일 동작함.
        if EXHIBITION_MODE:
            pixmap_mode = QPixmap(f'{IMG_PATH}{IMG_MODE_EXHIBITION}')
        else:
            pixmap_mode = QPixmap(f'{IMG_PATH}{IMG_MODE_SCHOOL}')
        pixmap_mode = pixmap_mode.scaled(self.lb_mode_img.size().width(), self.lb_mode_img.size().height(), Qt.KeepAspectRatio) 
        self.lb_mode_img.setPixmap(pixmap_mode)


    def createHandViewWidget(self):
        #################################################
        # [중요] 위젯에 다른 위젯 넣기
        #################################################
        self.handCountView = HandCountView()
        self.handCountView.setParent(self.fr_main_quiz)
        self.handCountView.lower()
        self.handCountView.hide()   # 손가락 레이블
        # self.handCountView.show()   # 손가락 레이블

        # self.handView.hide()     # 초기화면 숨기기

    # 손가락 / 터치 모드 이미지 표시
    # mode = None, fingerCount, thumbUpDown, touchMode
    def viewInputModImg(self, mode):
        print(' < 1 >>> finger count 5 display = ', mode)
        if mode == None:
            self.lb_inputModeImg.hide()
            return
        elif mode == 'fingerCount1':
            pixmap = QPixmap(f'{IMG_PATH}{IMG_MODE_FINGER_COUNT1_FILE}')
            self.lb_inputModeImg.setAlignment(Qt.AlignCenter)
        elif mode == 'fingerCount2':
            pixmap = QPixmap(f'{IMG_PATH}{IMG_MODE_FINGER_COUNT2_FILE}')
            self.lb_inputModeImg.setAlignment(Qt.AlignCenter)
        elif mode == 'fingerCount3':
            pixmap = QPixmap(f'{IMG_PATH}{IMG_MODE_FINGER_COUNT3_FILE}')
            self.lb_inputModeImg.setAlignment(Qt.AlignCenter)
        elif mode == 'fingerCount4':
            pixmap = QPixmap(f'{IMG_PATH}{IMG_MODE_FINGER_COUNT4_FILE}')
            self.lb_inputModeImg.setAlignment(Qt.AlignCenter)
        elif mode == 'fingerCount5':
            pixmap = QPixmap(f'{IMG_PATH}{IMG_MODE_FINGER_COUNT5_FILE}')
            self.lb_inputModeImg.setAlignment(Qt.AlignCenter)
        elif mode == 'thumbUpDown':
            pixmap = QPixmap(f'{IMG_PATH}{IMG_MODE_FINTER_THUMB_FILE}')
            self.lb_inputModeImg.setAlignment(Qt.AlignCenter)
        elif mode == 'touchMode':
            pixmap = QPixmap(f'{IMG_PATH}{IMG_MODE_TOUCH}')
            self.lb_inputModeImg.setAlignment(Qt.AlignLeft)

        pixmap2 = pixmap.scaled(self.lb_inputModeImg.size().width(), self.lb_inputModeImg.size().height(), Qt.KeepAspectRatio) 
        self.lb_inputModeImg.setPixmap(pixmap2)
        self.lb_inputModeImg.show()


    def displayInt(self):
        # clear
        self.lb_no.setText('문제: ' )
        self.lb_code.setText('CODE: ' )
        self.lb_type.setText('영역: ')

        if not EXHIBITION_MODE:
            self.lb_usrName.setText(f'{val.st_id} {val.st_name}')
        else:
            self.lb_usrName.setText(f'{val.st_school} {val.st_name}')
        self.lb_timeView.setText(QTime.fromString(val.speedQuizTime, 'm:s').toString('mm:ss'))

        if QUIZ_TYPE == 'speed':
            self.lb_type2.setText('퀴즈 유형: ' + '스피드 퀴즈')
        elif QUIZ_TYPE == 'golden':
            self.lb_type2.setText('퀴즈 유형: ' + '골든벨 퀴즈')
    
        if Q_CODE_VIEW :
            self.lb_code.show()
        else:
            self.lb_code.hide()
        self.lb_correct.setText(str(val.st_right_cnt))
        # 오답/PASS
        self.lb_wrongpass.setText(f'{str(val.st_wrong_cnt)}/{str(val.st_pass_cnt)}')
        # 점수
        self.lb_score.setText(str(val.st_score))


    ##########################################################
    def quiz1_CreatePreView(self):
        self.lb_waitMsg.show()
        self.lb_question.setText('')
        self.lb_question.hide()
        self.lb_chAnswer.setText('')
        self.lb_chAnswer.hide()

        self.lb_question2.hide()
        self.lb_question2Img.hide()

        msg ='<html><head/><body><p align="center">문제를 생성하고 있습니다.</p><p align="center"> 잠시만 기다려주세요</p></body></html>'
        self.lb_waitMsg.setText(msg)
        # 준비(Ready) 시간동안
        #   문제 파일 로딩

        # QTimer.singleShot(300, lambda :self.quizFileLoad('quiz_random') )

        # 스레드 처리하고
        # 완료되면, quiz_reay.emit()
        # self.quizFileLoadSave('quiz_random')    # > 이미지 처리 시간 지연 발생 > Thread 처리
        

    def quiz3_ReadyCount_Start(self):
        self.lb_waitMsg.hide()
        #   기본 안내문
        text = f' <p> &nbsp;&nbsp;&nbsp; 스피드 퀴즈를 시작합니다. </p> \
                  <p> &nbsp;&nbsp;&nbsp; 제한 시간({val.speedQuizTime}) 안에 최대한 많은 답을 맞추어야 합니다. </p>\
                  <p> &nbsp;&nbsp;&nbsp; - 점수 계산 방법 </p> \
                  <p> &nbsp;&nbsp;&nbsp;&nbsp;&nbsp; 정답: +{Q_ANSWER_SCORE}점, 오답 : {Q_WRONG_SCORE}점, 통과(Pass): {Q_PASS_SCORE}점 </p>'
        text2 = QUIZ_FORMAT.replace('_quiz', text)  # 기본 양식안에 넣기
        #   안내문 파일이 있는 경우
        if self.msg_quiz_countdown != None:
            text = self.msg_quiz_countdown
            text2 = QUIZ_FORMAT.replace('_quiz', text)  # 기본 양식안에 넣기
            text2 = text2.replace('[time]', val.speedQuizTime).replace('[answer_score]', str(Q_ANSWER_SCORE)).replace('[wrong_score]', str( Q_WRONG_SCORE)).replace('[pass_score]', str(Q_PASS_SCORE))
        # 안내문 표시
        self.lb_question.setText(text2)

        # 도전자
        if not EXHIBITION_MODE:
            self.lb_usrName.setText(f'{val.st_id} {val.st_name}')
        else:
            self.lb_usrName.setText(f'{val.st_school} {val.st_name}')
        
        self.lb_question.show()
        self.lb_question2.hide()
        self.lb_question2Img.hide()
        self.lb_quizResult.show()   # 결과
        self.lb_chAnswer.hide()     # 1) ~ 5)
        self.lb_inputModeImg.hide() # 손가락 안내 이미지 감추기
        self.lb_test_hint.hide()

        # clear
        # self.lb_quizResult.setText('') 
        # 0 sec
        self.lb_quizResult.setStyleSheet("color: yellow; font: 700 200pt ;")
        self.lb_quizResult.setText('Ready')
        self.sound_play_signal.emit('Q_READY')
        # 1 sec
        QTimer.singleShot(1000, lambda : self.lb_quizResult.setStyleSheet("color: yellow; font: 700 300pt ;"))
        QTimer.singleShot(1000, lambda : self.lb_quizResult.setText('3'))
        # 2 sec
        QTimer.singleShot(2000, lambda : self.lb_quizResult.setText('2'))
        # 3 sec
        QTimer.singleShot(3000, lambda : self.lb_quizResult.setText('1'))
        # 4 sec
        QTimer.singleShot(4000, lambda : self.lb_quizResult.setText(''))
        # 퀴즈 시작
        QTimer.singleShot(4000, self.quiz4_Start)

        # if self.test_hint:
        #     self.lb_test_hint.show()
        # else:
        #     self.lb_test_hint.hide()


    ##########################################################
    def quiz4_Start(self):
        # self.quizFileLoad()
        self.quiz41_Init('start')
        self.quiz42_Next()
    ##########################################################

    ##########################################################
    def quiz4_Test_Start(self):
        # self.quizFileLoad()
        self.quiz41_Init('test')
        self.quiz42_Next()
    ##########################################################

    # def loadQuiz_init(self, file):
    def quiz2_FileLoadSave(self, random_seq_mode):
        # 1. 퀴즈파일 생성 로딩/ 랜덤화된 문제로 저장 / 문제 전처리
        # # self.xlsxQuizLoad.quizPreSave(val.quizFileNum, random_seq_mode)
        # self.quiz2_LoadSaveThread = QuizLoadSaveThread(random_seq_mode)
        # # self.quiz2_LoadSaveThread.finished.connect(self.quiz2_LoadSaveThread_finished_slot)
        
        # QThread 방법 1)
		# https://realpython.com/python-pyqt-qthread/

        # # Step 2: Create a QThread object
        # self.thread = QThread()
        
        # # Step 3: Create a worker objec
        # self.xlsxQuizLoad = XlsxQuizLoad(val.quizFileNum, QUIZ_SELECT_MODE)

        # # Step 4: Move worker to the thread
        # self.xlsxQuizLoad.moveToThread(self.thread)

        # # Step 5: Connect signals and slots
        # self.xlsxQuizLoad.finished.connect(self.quiz2_LoadReady_slot)

        # # Step 6: Start the thread
        # self.thread.start()

        self.xlsxQuizLoad = XlsxQuizLoad(val.quizFileNum, random_seq_mode)
        self.xlsxQuizLoad.finished.connect(self.quiz2_LoadReady_slot)
        print('*** 1 ****')
        self.xlsxQuizLoad.start()



    def quiz2_LoadReady_slot(self):
        print('*** 2 ****')
        # 2. 검증/랜덤화된 문제 로딩
        self.read_wb = load_workbook(f'{TEMP_PATH }{QUIZ_RANDOM_FILE}')
        self.read_ws = self.read_wb.sheetnames
        self.read_ws = self.read_wb[self.read_ws[0]]
        
        # self.image_loader = SheetImageLoader(self.read_ws)

        self.qNoMax = self.read_ws.max_row - QUIZ_USER_FILE_START_ROW # 최대 문항수
        print('최대 문항수 :', self.qNoMax )

        # 1. 사용자 퀴즈 데이터 생성
        self.user_quiz_data = []
        #   [R_INFO][0] = ID, 이름, 시작 시간, (풀이시간)
        val.st_quizStartTime = QDateTime.currentDateTime().toString("yyyy-MM-dd hh:mm:ss")
        self.user_quiz_data.append( [ val.st_id, val.st_name, val.st_quizStartTime ])  # [R_INFO = 0]
        #   [R_QCNT = 1; R_QNUM = 2; R_RES = 3; R_RWP = 4][0] = 이름 기록

        self.user_quiz_data.append( ['문제카운트'] )     # R_QCNT = 1
        self.user_quiz_data.append( ['퀴즈  번호'] )     # R_QNUM = 2
        self.user_quiz_data.append( ['정      답'] )     # R_ANS =3
        self.user_quiz_data.append( ['응      답'] )      # R_RES = 4
        self.user_quiz_data.append( ['채      점'] )      # R_RWP = 5
        self.user_quiz_data.append( ['응답시간'] )      # R_TIME = 6 ms
        self.user_quiz_data.append( ['입력방법'] )      # R_HOTO = 7 (T)터치/(C)카운터/(F)얼굴기울림
        self.user_quiz_data.append( ['문제반복횟수'] )  # R_QREP = 8 동일한 퀴즈 사용자 몇번째 풀이
        print(self.user_quiz_data)

        # 처음 퀴즈시작
        # self.quiz42_Next('golden_bell')
        # self.quizProgress('speed_quiz')

        # Signal : 준비완료 시그널 발생
        self.quiz2_loadReady_finished.emit()

    def quiz41_Init(self, mode):
        val.st_quiz_cnt = 0  # 문제생성 카운더 초기화

        self.lb_no.show()
        self.lb_type.show()
        
        # 그림자 효과 적용
        # self.lb_chAnswer.setStyleSheet("border : 8px solid black")
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(5)
        shadow.setOffset(3)
        shadow.setColor(Qt.white)
        self.lb_chAnswer.setGraphicsEffect(shadow)

        # 일반 문제
        shadow1 = QGraphicsDropShadowEffect(self)
        shadow1.setBlurRadius(5)
        shadow1.setOffset(2)
        shadow1.setColor(Qt.white)
        self.lb_question.setGraphicsEffect(shadow1)

        # 이미지 포함 문제
        shadow2 = QGraphicsDropShadowEffect(self)
        shadow2.setBlurRadius(5)
        shadow2.setOffset(2)
        shadow2.setColor(Qt.white)
        self.lb_question2.setGraphicsEffect(shadow2)

        self.keyOneInput = True   # 퀴즈 중 다른 키 입력되어 연속 동작 방지

        ##################################################################################################
        # Qtimer 스탑워치
        #   https://wikidocs.net/38522
        #   https://github.com/PeterJohnson/countdown/blob/master/countdown.pyw
        #   https://doc.qt.io/qtforpython-5/PySide2/QtCore/QTimer.html#PySide2.QtCore.PySide2.QtCore.QTimer.remainingTime
        # Qtime
        #   https://wikidocs.net/37460
        if mode == "start":
            if Q_TEST_HINT:
                self.test_hint = True
            else:
                self.test_hint = False

            self.timerSpeedQ = QTimer()
            self.timerSpeedQ_rem = QTime.fromString(val.speedQuizTime, 'm:s')
            self.timerSpeedQ.setInterval(QTIMER_INTERVAL)
            self.timerSpeedQ.timeout.connect(self.timerSpeedQ_timeout)
            self.timerSpeedQ.start()

        elif mode == "test":
            self.test_hint = True

        # 생성된 랜덤 퀴즈 파일에서 이미지 row 찾아 리스트로 만들기
        self.quiz_image_rows = self.getImageCellRows()
    

    def timerSpeedQ_timeout(self):
        # <주의> QTime 은 Qtime과 연산이 되지 않음
        #   self.timerSpeedRemainingTime = QTime.currentTime() - self.timerSpeedStartTime
        # 대신 addsec(), addMSec() 가능
        #   https://doc.qt.io/qtforpython-5/PySide2/QtCore/QTime.html#PySide2.QtCore.PySide2.QtCore.QTime.addMSecs
        self.timerSpeedQ_rem = self.timerSpeedQ_rem.addMSecs(-QTIMER_INTERVAL)
        self.lb_timeView.setText(self.timerSpeedQ_rem.toString('mm:ss'))
        
        # qtime 비교 연산
        #   https://doc.qt.io/qtforpython-6/PySide6/QtCore/QTime.html
        # print( self.timerSpeedQ_rem.__eq__(QTime.fromString('0:0', 'm:s')) or
        #       self.timerSpeedQ_rem.__ge__(QTime.fromString(SPEED_QUIZ_TIME, 'm:s')) )
        
        # SPEED_QUIZ_WARNING1_TIME
        # 남은 시간을 가지고 타이머 색상 변경하기
        if self.timerSpeedQ_rem.__le__(QTime.fromString(SPEED_QUIZ_WARNING2_TIME, 'm:s')):
            if self.timeWarning != 2:
                self.lb_timeView.setStyleSheet('background-color: rgb(255, 170, 127); color: rgb(170, 0, 127);')
                self.sound_play_signal.emit('stop')
                self.sound_play_signal.emit('ticktock2')
                self.timeWarning = 2

        elif self.timerSpeedQ_rem.__le__(QTime.fromString(SPEED_QUIZ_WARNING1_TIME, 'm:s')):
            if self.timeWarning != 1:
                self.lb_timeView.setStyleSheet('background-color: rgb(255, 170, 127); color: rgb(170, 0, 127);')
                self.sound_play_signal.emit('ticktock1')
                self.timeWarning = 1

        else:
            self.lb_timeView.setStyleSheet('color : white;')

        timeOverChk = ( self.timerSpeedQ_rem.__eq__(QTime.fromString('0:0', 'm:s')) or
              self.timerSpeedQ_rem.__ge__(QTime.fromString(val.speedQuizTime, 'm:s')) )

        if timeOverChk :
            self.lb_timeView.setStyleSheet('color : white;')    # 타이머 색상 복원
            self.timerSpeedQ.stop()
            print(" >> timer stop ")
            self.main_to_signal.emit("quiz_end")

    def timerSpeedQ_stop(self):
        print(' >> timerSpeedQ_stop : stop')
        try:
            self.timerSpeedQ.stop()
        except:
            pass

    def getImageCellRows(self):
        self.read_wb = load_workbook(f'{TEMP_PATH }{QUIZ_RANDOM_FILE}')
        self.read_ws = self.read_wb.sheetnames
        self.read_ws = self.read_wb[self.read_ws[0]]

    # getImageList(self)
        copy_sheet1_images = self.read_ws._images[:]
        _images = []
        for image in copy_sheet1_images:
            row = image.anchor._from.row 
            col = image.anchor._from.col
            cell = [row, col]
            if col == XLS_QUESTION_IMAGE:
                _images.append(cell)
        _images.sort(key=lambda x:x[0])
        _image_cells = [list(t) for t in set(tuple(item) for item in _images)]
        _image_cells.sort(key=lambda x:x[0])

        _rows = []
        for _item in _image_cells:
            _rows.append(_item[0]+1)
            
        self.read_wb.close()

        return _rows


    def quiz42_Next(self):
        # 다른키 입력 활성화
        self.keyOneInput = True

        # 틀릴때 까지 반복하는 모드
        val.st_quiz_cnt += 1
        print(f'DEBUG1 st_quiz_cnt : {val.st_quiz_cnt}')

        # 문제() 퀴즈 보여주기
        self.sound_play_signal.emit('Q_START')
        self.quizLoadNo(val.st_quiz_cnt)
            
        # 1번 문제 시작
        # 문제 종료후 결과 보이기

    def quizLoadNo(self, q_no):
        # 터치 연속 입력 제한
        self.key_repeat_ready = True

        # 문제 로딩
        # ###################################################################
        # ## 5. 엑셀파일 읽기
        # read_wb = load_workbook(f'{QUIZ_RANDOM_FILE}', data_only=True)

        # read_ws = read_wb.sheetnames     # 리스트 형태로 반환
        # read_ws = read_wb[read_ws[0]]
        
        # 퀴즈 row 라인 불려어기
        _row = self.read_ws[q_no + QUIZ_USER_FILE_START_ROW]   # tupe type

        if q_no > self.qNoMax -1:
            # 최대 문항 초과 퀴즈 종료
            q_no = self.qNoMax
            print(" 더 이상 문제가 없음 ")
            # val.st_quiz_cnt -= 1 # 문제가 없는 경우 퀴즈 카운터 원래대로

        self.quiz_line = []
        
        # 퀴즈 row 데이터 텍스트 검사, 
        for idx, cell in enumerate(_row):
            data = cell.value
            if data != None:
                if type(data) == type('s'):
                    # xlsx 파일 읽어 드릴때 공백문자 '\xa0' 가 발생하는 경우 처리
                    self.quiz_line.append(cell.value.replace('\xa0', ' '))
                else:
                    self.quiz_line.append(data)
            else:
                if idx < Q_CHOICE:  # 퀴즈이미지 부분은 None 인 경우가 있으므로 column을 유지하기 위함.
                    self.quiz_line.append(data)

        # print('debug: quiz line =', self.quiz_line)

        self.quiz_line[Q_ANSWER] =  str(self.quiz_line[Q_ANSWER]) # 정답 int -> str

        # quiz_line = [3, 'c', '유퀴즈', '다음 중 과일과 채소 중 ‘채소’에 속하는 것은 무엇일까요?', 1, '토마토', '아보카도', '감', None, None]
        # [TEST] 퀴즈 1문제 출력
        # print(self.quiz_line)

        ###################################################################
        # ??  displayInt() 로 대체 가능 여부 고려!
        # widget view
        self.lb_no.setText('문제: ' + str(val.st_quiz_cnt))
        self.lb_code.setText('CODE: ' + str(self.quiz_line[Q_NO]) )
        self.lb_type.setText('영역: '  + str(self.quiz_line[Q_SUBJECT]) )

        # 퀴즈 결과 숨기기
        self.lb_quizResult.hide()
        
        # 남은시간
        # ??

        # 정답
        self.lb_correct.setText(str(val.st_right_cnt))
        # 오답/PASS
        self.lb_wrongpass.setText(f'{str(val.st_wrong_cnt)}/{str(val.st_pass_cnt)}')
        # 점수
        self.lb_score.setText(str(val.st_score))

        # 문제 표시
        quiz_text = str(self.quiz_line[Q_QUESTION])

        # 문제가 없는 경우 메시지
        if quiz_text == 'None':
            quiz_text = ' 문제를 모두 풀이하였습니다. <br><br> 더 이상 준비된 문제가 없습니다.'

        ###### 6_text_underline.py 에서 테스트 적용 #########
        # 부정 문제 밑줄 넣기 : 최초 만나는 1개의 단어만 가능
        for text in UNDERLINE_WORDS:
            # cnt = cnt + quiz_text.count(text)
            length = len(text)

            idx = quiz_text.find(text)

            if idx > 0:
                quiz_text = quiz_text[0:idx] + '<u>' + quiz_text[idx:idx+length] +'</u>' + quiz_text[idx+length:]
        # print(quiz_text)
        ####################################################

        quiz_view = QUIZ_FORMAT.replace('_quiz', quiz_text)


        ####################################################
        # (문제표시) 이미지가 포함된 문항인지 확인하기
        # print(getImageCellRows_dic.keys())
        # print(self.quiz_line[Q_NO])

        #  row[0].row           # 현재 행의 첫 번째 셀을 기준으로 행 번호 가져오기
        print('CC _row[0].row = ', _row[0].row)
        if _row[0].row in self.quiz_image_rows:
            # 이미지가 포함된 경우 (image)
            self.lb_waitMsg.setText('')
            self.lb_waitMsg.hide()
            self.lb_question2.setText(quiz_view)
            self.lb_question.hide()
            self.lb_question2.show()

            # 이미지 로드 row_column = [COL_, ROW_]
            row_column = [q_no + QUIZ_USER_FILE_START_ROW, Q_QUESTION_IMAGE+1]
            column_letter = get_column_letter(row_column[COL_])
            cell_index = f'{column_letter}{row_column[ROW_]}'   # A6 같은 형태로 변환
            print('이미지 셀 :', cell_index)
            
            # try:
            # PIL Image to QPixmap conversion issue
            #   https://stackoverflow.com/questions/34697559/pil-image-to-qpixmap-conversion-issue
            #   같은 이미지를 다시 로딩하면 문제가 발생함.

            ##################################################################
            # ERROR 1 - 해결 방법: 파일을 닫고, 다시 이미지 로딩시 문제 해결 됨.
            self.read_wb.close()
            self.read_wb = load_workbook(f'{TEMP_PATH }{QUIZ_RANDOM_FILE}')
            self.read_ws = self.read_wb.sheetnames
            self.read_ws = self.read_wb[self.read_ws[0]]
            ##################################################################

            image_loader = SheetImageLoader(self.read_ws)
            img = image_loader.get(cell_index)      # ERROR 1 - 다시 이미지 로드하는 경우 에러
            
            qimg = ImageQt(img)
            pix = QtGui.QPixmap.fromImage(qimg)
            self.lb_question2Img.setPixmap(pix)
            self.lb_question2Img.show()

            # print('이미지 셀:', cell_index, ' 타입:', type(img))
            # except:
            # self.lb_question2Img.hide()

        else:
            # 이미지가 미 포함된 경우 (no image)
            self.lb_waitMsg.setText('')
            self.lb_waitMsg.hide()
            self.lb_question.setText(quiz_view)
            self.lb_question.show()
            self.lb_question2.hide()
            self.lb_question2Img.hide()
        
        ####################################################
        # 문제 출력후 응답하는데 걸리는 시간 계산을 위함.
        self.quiz_start_time = QTime.currentTime()

        ####################################################
        # (1~5 선택 답안 표시)
        self.lb_chAnswer.hide()
        choice = []
        if self.quiz_line[Q_TYPE] == Q_TYPE_CHOICE or self.quiz_line[Q_TYPE] == Q_TYPE_MIX:
            
            # choice = [self.quiz_line[Q_CHOICE], self.quiz_line[Q_CHOICE+1], self.quiz_line[Q_CHOICE+2], self.quiz_line[Q_CHOICE+3], self.quiz_line[Q_CHOICE+4]]
            for c in range(5):
                try:
                    choice.append(self.quiz_line[Q_CHOICE+c])
                except IndexError:
                    pass
            
            # print('choice : ', choice)
            i = 0
            for j in range(len(choice)):
                if choice[j] != None:
                    i += 1
            # print('choice count : ', i)
                    
            if FINGER_COUNT_INPUT_MODE:
                # self.viewInputModImg(f'fingerCount{i}')
                self.main_to_signal.emit(f'fingerCount{i}')
                print(f' < 2 >>> 다음문제 로드 self.main_to_signal.emit  =  "fingerCount{i}"')
            else:
                self.viewInputModImg(None)
                self.main_to_signal.emit(f'button_choice{i}')


            ch = ['', '', '', '', '']
            for i  in range(5):
                try:
                    ch[i]= f'{i+1}) '+ str(self.quiz_line[Q_CHOICE+i])
                    # print( str(self.quiz_line[Q_CHOICE+i]) )
                except IndexError:
                    pass

            # ans_view = SELECT_FORMAT.replace('_ch1', ch1).replace('_ch2', ch2).replace('_ch3', ch3).replace('_ch4', ch4).replace('_ch5', ch5)
            ans_view = SELECT_FORMAT.replace('_ch1', ch[0]).replace('_ch2', ch[1]).replace('_ch3', ch[2]).replace('_ch4', ch[3]).replace('_ch5', ch[4])
            self.lb_chAnswer.setText(ans_view)
            self.lb_chAnswer.show()

        # OX 버튼 / 하단 숨기기    
        elif self.quiz_line[Q_TYPE] == Q_TYPE_OX:

            if FINGER_COUNT_INPUT_MODE:
                self.viewInputModImg('thumbUpDown')
                self.main_to_signal.emit('thumbUpDown')
            else:
                self.viewInputModImg('touchMode')
                self.main_to_signal.emit('button_ox')

            self.lb_chAnswer.hide()

        # 문제가 없는 경우
        else:
            self.viewInputModImg(None)
            self.main_to_signal.emit('button_clear')

        # 테스트 모드시 정답 표시
        if self.test_hint:
            self.lb_test_hint.show()
            q_answer = self.quiz_line[Q_ANSWER]
            ansText = ''
            if q_answer == 'O':
                ansText = ansText + '.'
            elif q_answer == 'X':
                ansText = ansText + '..'
            else:
                q_answer = int(q_answer)
                for i in range(q_answer):
                    ansText = ansText + '.'
            self.lb_test_hint.setText(ansText)
        else:
            self.lb_test_hint.hide()




    def closeEvent(self, event):
        self.thread.stop()
        event.accept()


    @Slot(np.ndarray)
    def update_image_slot(self, qt_img):
        """ X:주의 슬롯에 넣으면, 처리 못함.가상 키보드 """
        # self.vr_keyboard(cv_img)

        """Updates the image_label with a new opencv image"""
        # qt_img = self.convert_cv2qt(cv_img) -> thread 에서 처리하는 것으로 변경함.
        self.lb_imageView.setPixmap(qt_img)
    

    @Slot(str)
    def key_one_slot(self, key):
        if key == "OK" or key == "QUIT":
            return
        if self.keyOneInput:
            self.keyOneInput = False   # 퀴즈 중 다른 키 입력금지
            use_key =['1','2','3','4','5','O','X',"PASS"]
            if key in use_key:
                # self.sounds[self.soundDict[key]].play()
                self.sound_play_signal.emit(key)
                # 손가락 터치 응답을 처리
                QTimer.singleShot(QUIZ_RESULT_DELAY_TIME, lambda: self.quizResult(key))


    def quizResult(self, key):
        # 손가락 터치 응답/ 손가락 카운터 응답을 모두 처리
        quiz_end_time = QTime.currentTime()
        quiz_elapsed_time = self.quiz_start_time.msecsTo(quiz_end_time)

        print(f' >>> quizResult key = {key} , elapsed_time = {quiz_elapsed_time} ms')
        if key == 'PASS':
            # PASS 표시
            self.sound_play_signal.emit('Q_PASS')
            self.lb_quizResult.setStyleSheet("color: rgb(239, 176, 49); font: 700 200pt ;")
            self.lb_quizResult.setText('PASS')
            self.lb_quizResult.show()

            _r_rwp = 'Pass' # 패스
            val.st_pass_cnt += 1
            val.st_score += Q_PASS_SCORE
            # 다음 문제로
            QTimer.singleShot(QUIZ_RESULT_VIEW_TIME, self.quiz42_Next)
            
        # 정답 선택의 경우 
        elif key == self.quiz_line[Q_ANSWER]:
            self.sound_play_signal.emit('Q_CORRECT')
            self.lb_quizResult.setStyleSheet("color: blue; font: 700 300pt ;")
            self.lb_quizResult.setText('O')
            self.lb_quizResult.show()

            _r_rwp = 'Right' # 정답
            val.st_right_cnt += 1
            val.st_score += Q_ANSWER_SCORE
            # 다음 문제로
            QTimer.singleShot(QUIZ_RESULT_VIEW_TIME, self.quiz42_Next)

        # 오답 선택의 경우
        else:
            self.sound_play_signal.emit('Q_WRONG')
            self.lb_quizResult.setStyleSheet("color: red; font: 700 300pt ;")
            self.lb_quizResult.setText('X')
            self.lb_quizResult.show()

            _r_rwp = 'Wrong' # 오답
            val.st_wrong_cnt += 1
            val.st_score += Q_WRONG_SCORE

            # 다음 문제로
            QTimer.singleShot(QUIZ_RESULT_VIEW_TIME, self.quiz42_Next)
        
        # 결과 처리
        if FINGER_COUNT_INPUT_MODE == True:
            _how_to_respond = 'Count'   # finger count input
        else:
            _how_to_respond = 'Touch'   # finger touch input

        # How many times to repeat the quiz
        ''' ################################################################## '''
        ''' ## 퀴즈 결과를 데이터베이스(sqlite DB)에 기록                     ## '''
        # 손님이 아닌 경우.
        if val.st_id != '':
            ######## 데이터 베이스 연결 ########
            self.conn = sqlite3.connect(f'{DB_PATH}{DB_BASE_FILE}') 
            # cursor 생성
            self.cursor = self.conn.cursor()  # db 처음 시작 지점
            quiz_file = val.quizFileList[val.quizFileNum]

            # 1) user_id 가 기존에 문제를 풀이한 기록이 있는지 DB에서 찾기
                              # answer_id,   user_id,   quiz_id, file_name,  quiz_number             , answer, answer_time, how_to, quiz_repeat, date_time
            r_answerData = AnswerData(None, int(val.st_id), None, quiz_file, int(self.quiz_line[Q_NO]) )
            answerDb = AnswerDAO(self.conn, quiz_file)
            read_repeat = answerDb.read_repeat(r_answerData)

            # 2) user_id 문제 풀이 결과를 DB에 기록
            read_repeat += 1
                            # answer_id,   user_id,   quiz_id, file_name,  quiz_number              , answer, answer_time,      how_to, quiz_repeat, date_time
            answerData = AnswerData(None, int(val.st_id), None, quiz_file, int(self.quiz_line[Q_NO]), _r_rwp, quiz_elapsed_time, _how_to_respond, read_repeat )
            answerDb.insert(answerData)

            ######## 연결 해제 ########
            self.cursor.close()
            self.conn.close()
        else:
            read_repeat = 0
        ''' ################################################################## '''
        # db 작업 완료후, db 에서 read_repeat 조회하기
        
        self.user_quiz_data[R_QCNT].append(val.st_quiz_cnt)              # 문제카운트
        self.user_quiz_data[R_QNUM].append(self.quiz_line[Q_NO])         # 퀴즈  번호
        self.user_quiz_data[R_ANS].append(self.quiz_line[Q_ANSWER])      # 정      답
        self.user_quiz_data[R_RES].append(key)                           # 응      답
        self.user_quiz_data[R_RWP].append(_r_rwp)             # 채점 Right / Wrong / Pass
        self.user_quiz_data[R_TIME].append(quiz_elapsed_time) # 응답시간 ms
        self.user_quiz_data[R_HOTO].append(_how_to_respond)   # 입력방법: (Touch)터치/(Count)카운터/(F)얼굴기울림
        self.user_quiz_data[R_QREP].append(read_repeat)       # db 조회 몇번째 풀이인지

        for i in range(len(self.user_quiz_data)):
            print( self.user_quiz_data[i] )



        

    ####################### 수정 필요 ##############################
    @Slot(str)
    def finger_repeat_slot(self, finger):
        # print('Slot finger = ', finger)
        use_finger = ['1','2','3','4','5','O','X']
        if finger in use_finger:
            # self.sound_play_signal.emit(finger)
            # print('m_quiz Slot finger_repeat_slot = ', finger)
            self.handCountView.lb_fingerText.setText(str(finger))

            if self.FingerKeyTimer.value(finger):
                self.sound_play_signal.emit(finger)
                # self.main_to_signal.emit(finger)
                print('**** Quiz Result = ', finger)
                # 손가락 갯수 카운터 응답을 처리
                QTimer.singleShot(QUIZ_RESULT_DELAY_TIME, lambda: self.quizResult(finger))
    
    @Slot(list)
    def hand_position_slot(self, handPosition):
        X = 0; Y = 1; LENGTH = 2
        if len(handPosition) == 0:
            self.handCountView.hide()
        else:
            self.handCountView.show()
            x = handPosition[X]
            y = handPosition[Y]
            length = handPosition[LENGTH]

            radio = length/100
            self.handCountView.setZoom(radio)

            self.handCountView.setMove(x, y, radio)

    ###############################################################

    @Slot(str)
    def key_repeat_slot(self, key):
        if self.key_repeat_ready == False:
            return
        if key == "QUIT" :
            keyState = True
        else:
            keyState = False

        # 누르는 시간(KEY3_PUSH_TIME)을 넘는 경우
        if self.PressKeyTimer.state(keyState):
            self.sound_play_signal.emit("QUIT")
            self.main_to_signal.emit("quiz_end")    # 메인 에 신호를 보낸다.
            self.self.key_repeat_ready = False
    
    @Slot(str)
    def fps_signal_slot(self, fps_str):
        self.lb_fps.setText(fps_str)

#######################################################################################
# 데이터 베이스 연결
#######################################################################################
class AnswerData:
    def __init__(self, answer_id=None, user_id=None, quiz_id=None, file_name=None, quiz_number=None, answer=None, answer_time=None, how_to=None, quiz_repeat=None, date_time=None):
        self.answer_id = answer_id      # 0
        self.user_id = user_id          # 1
        self.quiz_id = quiz_id          # 2
        self.file_name = file_name      # 3
        self.quiz_number = quiz_number  # 4
        self.answer = answer            # 5
        self.answer_time = answer_time  # 6
        self.how_to = how_to            # 7
        self.quiz_repeat = quiz_repeat  # 8
        self.date_time = date_time      # 9
    def __str__(self):
        return f"AnswerData(answer_id={self.answer_id}, user_id={self.user_id}, quiz_id={self.quiz_id}, file_name={self.file_name}, quiz_number={self.quiz_number}, answer={self.answer}, answer_time={self.answer_time}, how_to={self.how_to}, quiz_repeat={self.quiz_repeat}, date_time={self.date_time})"

class AnswerDAO:
    def __init__(self, conn, file_name):
        self.conn = conn
        self.cursor = conn.cursor()
        self.file_name = file_name

    def insert(self, vo):
        _datetime = QDateTime.currentDateTime().toString('yyyy-MM-dd hh:mm:ss')
        try:
            self.cursor.execute("INSERT INTO answer(user_id, file_name, quiz_number, answer, answer_time, how_to, quiz_repeat, date_time) VALUES(?, ?, ?, ?, ?, ?, ?, ?)",
                (vo.user_id, vo.file_name, vo.quiz_number, vo.answer, vo.answer_time, vo.how_to, vo.quiz_repeat, _datetime))
            self.conn.commit()
        except Exception as e:
            print(f'ERR answer insert : {e}')

    def read_repeat(self, vo):
        try:
            self.cursor.execute("SELECT * FROM answer WHERE user_id=? AND file_name=? AND quiz_number=?",
                                 (vo.user_id, vo.file_name, vo.quiz_number))
            rows = self.cursor.fetchall()
            return len(rows)
        except Exception as e:
            print(f'ERR read_repeat : {e}')
            return 0

    '''
    m_quiz.py
    동일 문제 반복 횟수 db 저장후 조회하여 보기
    _quiz_repeat = 0
    '''


#######################################################################
# 퀴즈 텍스트 형식 지정 | QUIZ_FORMAT.replace('_quiz', '안녕하세요')

QUIZ_FORMAT = '<style type=text/css> \
p.margin { \
    margin-top: 30px; \
    margin-bottom: 30px; \
    margin-right: 30px; \
    margin-left: 30px; \
    line-height: 110%; \
} \
</style> \
<p class="margin">_quiz</p>'

#######################################################################
# 선택 문항 형식 지정 | SELECT_FORMAT.replace('_ch1', ch1).replace('_ch2', ch2).replace('_ch3', ch3).replace('_ch4', ch4).replace('_ch5', ch5)

SELECT_FORMAT = ' <style type=text/css> \
p.margin { \
    margin-top: 15px; \
    margin-bottom: 15px; \
    margin-right: 30px; \
    margin-left: 30px; \
    line-height: 110%; \
} \
</style> \
<p class="margin">_ch1</p> \
<p class="margin">_ch2</p> \
<p class="margin">_ch3</p> \
<p class="margin">_ch4</p> \
<p class="margin">_ch5</p> '

#######################################################################

if __name__=="__main__":
    from _main import *
    PySide6Ui('ui_quiz.ui').toPy()

    app = QApplication(sys.argv)
    # _main 의 객체
    a = MyWidonws()

    # # TEST
    # quiz = QuizDisplay()
    # quiz.quizLoadNo(1)

    ## 카운트 다운 후 퀴즈 테스트 ##
    a.display_Quiz()
    QTimer.singleShot( 1600, a.threadVideo.button_clear)


    
    a.show()


    sys.exit(app.exec())








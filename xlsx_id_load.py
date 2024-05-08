'''
파이썬에서 엑셀을 사용하는 3가지 방법
https://blog.naver.com/550sn/222426744373
    Excel을 실행하고, 그 프로그램 내에서 명령을 하는 방식 (pywin32)
    Excel 파일에 직접 접근하여 원하는 형태로 수정하는 방식(openpyxl)
    컴퓨터가 쉽게 인식하는 변수 형태로 변환한 다음에 수정하는 방식(pandas).

    https://ponyozzang.tistory.com/664
https://wikidocs.net/91661
https://wooiljeong.github.io/python/openpyxl-tutorial/

https://pypi.org/project/openpyxl/
https://openpyxl.readthedocs.io/en/stable/

https://doitnow-man.tistory.com/entry/python-17-python으로-엑셀-다루기

Dictionary 값 수정, 추가, 삭제
    https://devinside.tistory.com/161
    

'''
from openpyxl import Workbook, load_workbook
from setup import *
import sqlite3

# 사용자 등록용 xlsx 파일에서 데이터 컬럼 위치 지정
GRADE = 0
CLASS = 1
NUMBER = 2
NAME = 3
GENDER = 4
MIN_ROW = 2 # 데이터 시작 줄

# temp_id.xlsx 파일 생성, preSave_xlsx_list=[] 리스트의 인덱스
#   -> def xlsx_id_read(self) 에서 위치가 결정됨
# TO_ID = 0
# TO_GRADE = 1
# TO_CLASS = 2
# TO_NUMBER = 3
# TO_NAME = 4
# TO_GENDER = 5
TO_ID = 0; TO_SCHOOL =1;  TO_GRADE = 2; TO_CLASS = 3; TO_NUMBER = 4; TO_NAME = 5; TO_GENDER = 6; TO_ETC = 7
# preSave_xlsx_list = [[TO_ID, TO_GRADE, TO_CLASS, TO_NUMBER, TO_NAME, TO_GENDER]]


class IdLoad():
    def __init__(self):
        ##################################################################
        # 네이밍 규칙 : 7가지 꼭 지켜야할 네이밍 규칙  https://killu.tistory.com/52
        # DB schema 설계시 주의할 점 2가지 https://killu.tistory.com/51?category=895657
        # https://velog.io/@ohsol/MySQL-DB-명명-규칙-정리
        self.conn = sqlite3.connect(f'{DB_PATH}{DB_BASE_FILE}') 
        # cursor 생성
        self.cursor = self.conn.cursor()  # db 처음 시작 지점


    def init(self):
        self.db_create_users_Table()

        if not RANKING_EXHIBITION_MODE:

            # 일반 퀴즈 모드인 경우
            ##################################################################
            # 1. xlsx 에서 (교사, 학생)사용자 정보 가져오기
            ##################################################################
            #    id_dicList = 퀴즈 사용자 검색용 딕션너리, preSave_xlsx_list = db 에 사용할 데이터
            self.id_dicList, self.preSave_xlsx_list = self.xlsx_id_read()
            # print(preSave_xlsx_list)

            ##################################################################
            # 2. db.users table 내용 읽어오기
            ##################################################################
            db_id_load = self.db_read_All()

            # db_create_User(["3201", 3, 2, 1 ,"홍일남", "남"])
            self.compare_xls_db_id(self.preSave_xlsx_list, db_id_load)

            # self.cursor.close()
            # self.conn.close()

        else:
            ##################################################################
            # 전시회 퀴즈 모드인 경우
            ##################################################################
            self.db_create_User(["5555", '고실중',3, 2, 12 ,"홍길동", "남", ""])
            self.db_create_User(["6666", '고실중',3, 2, 11 ,"홍길순", "남", ""])
            self.db_create_User(["7777", '고실중',3, 2, 19 ,"이순신", "남", ""])
            self.conn.commit()
            # self.cursor.close()
            # self.conn.close()
            pass



    def gender_type(self, gender):
        if gender == '남자' or gender == '남성': gender = '남'
        elif gender == '여자' or gender == '여성': gender = '여'
        if not (gender == '남' or gender == '여'):
            gender = ''
        return gender

    def xlsx_id_read(self):
        ##################################################################
        # 학생id 엑셀파일 읽기
        ##################################################################
        # read_wb = load_workbook("student_id.xlsx", data_only=True)
        read_wb = load_workbook(F'{ID_PATH}{ST_ID_FILE}', data_only=True)

        read_ws = read_wb.sheetnames                    # 리스트 형태로 반환
        # sheet = workbook.get_sheet_by_name('Sheet1')    # 비추천 에러 발생
        read_ws = read_wb[read_ws[0]]

        # 엑셀파일 쓰기
        write_wb = Workbook()
        # 이름이 있는 시트를 생성
        write_ws = write_wb.active

        preSave_xlsx_list =[]       # 리스트 - db 데이터 입력 및 비교용
        id_dicList ={}              # 딕션너리 - 퀴즈에서 사용자 검색용
        for row in read_ws.iter_rows(min_row = MIN_ROW):
            if (row[GRADE].value != None) and (row[CLASS].value != None) \
                and (row[NUMBER].value != None) and (row[NAME].value != None):
                # print(row[GENDER].value)
                # print(str(row[2].value).zfill(2))

                # id 생성하기
                id_str = str(row[GRADE].value) + str(row[CLASS].value) + str(row[NUMBER].value).zfill(2)
                gender = self.gender_type(row[GENDER].value)
                                # 학번(문자), 학년, 반, 번호, 이름(문자), 성별
                                #
                # xlsx 파일에 기록하기
                #                TO_ID = 0; TO_SCHOOL =1;  TO_GRADE = 2; TO_CLASS = 3; TO_NUMBER = 4; TO_NAME = 5; TO_GENDER = 6; TO_ETC = 7
                write_ws.append([id_str, "", row[GRADE].value, row[CLASS].value, row[NUMBER].value, row[NAME].value, gender, ""])
                
                id_int = self.covert_noneInt(id_str)
                grade_int = self.covert_noneInt(row[GRADE].value)
                class_int = self.covert_noneInt(row[CLASS].value)
                number_int = self.covert_noneInt(row[NUMBER].value)
                name_str = row[NAME].value

                preSave_xlsx_list.append([id_int, "", grade_int, class_int, number_int, name_str, gender, ""])
                # print(id, row[NAME].value)

                            # *학번(문자), *이름(문자)
                id_dicList[id_str] = ["", row[GRADE].value, row[CLASS].value, row[NUMBER].value, row[NAME].value, gender, ""]

        ##################################################################
        # 교사id 엑셀 파일 읽기
        ##################################################################
        try:
            read_wb = load_workbook(F'{ID_PATH}{TE_ID_FILE}', data_only=True)

            read_ws = read_wb.sheetnames                    # 리스트 형태로 반환
            # sheet = workbook.get_sheet_by_name('Sheet1')    # 비추천 에러 발생
            read_ws = read_wb[read_ws[0]]

            # 엑셀파일 쓰기
            # write_wb = Workbook()
            # 이름이 있는 시트를 생성
            # write_ws = write_wb.active

            # id_dicList ={}
            for row in read_ws.iter_rows(min_row = MIN_ROW):
                if (row[NUMBER].value != None) and (row[NAME].value != None):
                    # print(row[3].value)
                    # print(str(row[2].value).zfill(2))
                    if row[CLASS].value != None:
                        id_str = str(row[CLASS].value) + str(row[NUMBER].value).zfill(3)
                    else:
                        id_str = str(row[NUMBER].value).zfill(3)
                                # 학번(문자), 이름(문자)
                    # write_ws.append([id_str, row[NAME].value])
                    gender = self.gender_type(row[GENDER].value)
                    # xlsx 파일에 기록하기
                    #                TO_ID = 0; TO_SCHOOL =1;  TO_GRADE = 2; TO_CLASS = 3; TO_NUMBER = 4; TO_NAME = 5; TO_GENDER = 6; TO_ETC = 7
                    id_int = self.covert_noneInt(id_str)
                    grade_int = self.covert_noneInt(row[GRADE].value)
                    class_int = self.covert_noneInt(row[CLASS].value)
                    number_int = self.covert_noneInt(row[NUMBER].value)
                    name_str = row[NAME].value

                    write_ws.append([id_str, "", grade_int, class_int, number_int, name_str, gender, ""])
                    preSave_xlsx_list.append([id_str, "", grade_int, class_int, number_int, name_str, gender, ""])
                    # print(id, row[NAME].value)
                                # 학번(문자), 이름(문자)
                    id_dicList[id_str] = ["", row[GRADE].value, row[CLASS].value, row[NUMBER].value, row[NAME].value, gender, ""]
        except FileNotFoundError:
            pass
        
        # 임시 파일로 저장
        write_wb.save(f"{TEMP_PATH}{TEMP_ID_FILE}")
        print(f'xlsx id load : {len(id_dicList)}명의 ID 정보를 로드 했습니다.')

        # 닫기
        read_wb.close()
        write_wb.close()


        return id_dicList, preSave_xlsx_list
    

    def covert_noneInt(self, _str):
        if _str == '' or _str == None:
            return None
        else:
            return int(_str)

    def db_create_users_Table(self):
        try:
            self.cursor.execute("""
                CREATE TABLE IF NOT EXISTS user(
                    id INTEGER PRIMARY KEY,
                    school TEXT,
                    grade INTEGER,
                    class INTEGER,
                    number INTEGER,
                    name TEXT NOT NULL,
                    gender TEXT,
                    visit_count INTEGER,
                    crete_date TEXT,
                    etc TEXT
                )
            """)
            print(f"[DB1] create users table OK!")
        
        except Exception as e:
            print(f"[DB1 Err] {e}")

    def db_create_User(self, data):
        CreteDate = QDate.currentDate().toString('yyyy-MM-dd')
        visit_count = 0

        # UserID=int(data[TO_ID]); School=data[TO_SCHOOL]; Grade=int(data[TO_GRADE]); Class=int(data[TO_CLASS]); Number=int(data[TO_NUMBER]); Name=data[TO_NAME]; Gender=data[TO_GENDER]; Etc=data[TO_ETC]
        UserID=data[TO_ID]; School=data[TO_SCHOOL]; Grade=data[TO_GRADE]; Class=data[TO_CLASS]; Number=data[TO_NUMBER]; Name=data[TO_NAME]; Gender=data[TO_GENDER]; Etc=data[TO_ETC]
        
        Gender = self.gender_type(Gender)
        try:
            self.cursor.execute("INSERT INTO user(id, school, grade, class, number, name, gender, visit_count, crete_date, etc) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)" ,
                        (UserID, School,Grade, Class, Number, Name, Gender, visit_count, CreteDate, Etc) )
            # conn.commit()
            # print(f"[DB2] db_create_User ({UserID}, {Grade}, {Class}, {Number}, {Name}, {Gender}, {ParticipationCount}, {CreteDate})")

        except Exception as e:
            print(f"[DB2 db_create_User Err] {e}")

    def db_update_User(self, data):
        CreteDate = QDate.currentDate().toString('yyyy-dd-MM')
        UserID=data[TO_ID]; School=data[TO_SCHOOL]; Grade=data[TO_GRADE]; Class=data[TO_CLASS]; Number=data[TO_NUMBER]; Name=data[TO_NAME]; Gender=data[TO_GENDER]; Etc=data[TO_ETC]
        Gender = self.gender_type(Gender)
        try:
            self.cursor.execute("UPDATE user SET grade=?, school=?, class=?, number=?, name=?, gender=?, crete_date=?, etc=? WHERE id=?",
                        (Grade, School, Class, Number, Name, Gender, CreteDate, Etc, UserID ) )
            # conn.commit()
            # print(f"[DB5] db_update_User ({UserID}, {Grade}, {Class}, {Number}, {Name}, {Gender}, {CreteDate})")
        except Exception as e:
            print(f"[DB5 Err] {e}")


    def db_find_UserID(self, UserID):
        try:
            self.cursor.execute("SELECT * FROM user WHERE id=?",(UserID,))
            ret = self.cursor.fetchall()
            if len(ret):
                _ret = list(ret[0]) #첫번째 값만 튜플 -> 리스트로 변환하여 리컨 
                return _ret
            else:
                return ret
        except Exception as e:
            print(f"[DB6 Err] {e}")

    def db_read_All(self):
        result = []
        try:
            self.cursor.execute("SELECT * FROM user")
            rows = self.cursor.fetchall()
            for row in rows:
                result.append(list(row)) 
            # print(f"[DB4 read_All] {result}")
            return result
        except Exception as e:
            print(f"[DB4 Err] {e}")

    def db_read_User(self, UserID):
        try:
            self.cursor.execute("SELECT * FROM user WHERE id=?", (UserID,))
            rows = self.cursor.fetchall()
            # print(f'[DB3] read {rows}')
            return rows
        except Exception as e:
            print(f"[DB3 Err] {e}")
            return []
        
    def db_delete_User(self, UserID):
        try:
            self.cursor.execute("DELETE FROM user WHERE id=?", (UserID,))
            # conn.commit()
            # print(f'[DB4] DELETE {UserID}')
        except Exception as e:
            print(f"[DB4 Err] {e}")


    ##################################################################
    # xls 파일이 기준으로 db user 테이브 업데이트 하기
    # 1. xls 에 있고, db 에 있으면, 비교
    #  1-1)  xls 에 있고, db 나머지가 다르면, db update

    #  1-2)  xls 에 있고, db 에 없으면, db insert

    # 2) db 에 있고, xls 없으면, db 삭제
    #   1.의 과정을 거치면서 list_db_mark 하기, (db 결과: 튜플 -> 리스트로 변환)
    #       list_db_marker 가 False 이면, 마지막에 해당 db delete
    ##################################################################

    def compare_xls_db_id(self, preSave_xlsx_list, db_id_load):
        # print(f' preSave_xlsx_list = {preSave_xlsx_list}')
        # print(f' db_id_load = {db_id_load}')
        # print(f' preSave_xlsx_list[7:8] = {preSave_xlsx_list[0][7:8]}')
        # print(f' db_id_load[9:10] = {db_id_load[0][9:10]}')

        db_id_load_id_list = []
        for in_user_db in db_id_load:
            db_id_load_id_list.append(in_user_db[TO_ID])

        user_xls_id_list = []
        for in_user_xls in preSave_xlsx_list:
            user_xls_id_list.append(in_user_xls[TO_ID])
        
        create_cnt = 0; delete_cnt = 0; update_cnt = 0

        def user_db_find(UserID):
            if UserID in db_id_load_id_list:
                return True
            else:
                return False
        
        def user_xls_find(UserID):
            if UserID in user_xls_id_list:
                return True
            else:
                return False

        # https://www.acmicpc.net/board/view/29708
        for in_user_xls in preSave_xlsx_list:
            # 1. xls 에 있고
            for in_db_id_load in db_id_load:
                compare = None
                if in_user_xls[TO_ID] == in_db_id_load[TO_ID]:
                    # 1. xls 에 있고, db 에 있으면, 비교
                    #   preSave_xlsx_list = [[2101, '', 2, 1, 1, '강대성', '남', '']
                    #   db_id_load        = [[1101, 's', 1, 1, 1, '강진주', ' ', 0, '2024-29-04', '']
                    if  in_user_xls[:7] == in_db_id_load[:7] and in_user_xls[7:8] == in_db_id_load[9:10]:
                        compare = True
                    else:
                        compare = False
                #  1-1)  xls 에 있고, db 나머지가 다르면, db update
                if compare == False:
                    self.db_update_User(in_user_xls)
                    update_cnt += 1
            #  1-2)  xls 에 있고, db 에 없으면, db insert
            if user_db_find(in_user_xls[TO_ID]) == False:
                self.db_create_User(in_user_xls)
                create_cnt += 1

        for in_db_id_load in db_id_load:
            # 2) db 에 있고, xls 없으면, db 삭제
            if user_xls_find(in_db_id_load[TO_ID]) == False:
                self.db_delete_User(in_db_id_load[TO_ID])
                delete_cnt +=1

        self.conn.commit() 
        print(f'[db] user 데이터 베이스 : 추가 = {create_cnt} 삭제 = {delete_cnt} 수정 = {update_cnt}')

# VO(Value Object) : 값을 표현하는 클래스     
class DbUserData:
    def __init__(self, _id=None, _school=None, _grade=None, _class=None, _number=None, _name=None, _gender=None, _visit_count=None, _crete_date=None, _etc=None):
        self._id =_id
        self._school = _school
        self._grade = _grade
        self._class =_class
        self._number = _number
        self._name = _name
        self._gender = _gender
        self._visit_count =_visit_count
        self._crete_date = _crete_date
        self._etc = _etc
    def __str(self):
        return f"DbUserData(_id={self._id}, _school={self._school}, _grade={self._grade}, _class={self._class}, _number={self._number}, _name={self._name}, _gender={self._gender}, _visit_count={self._visit_count}, _crete_date={self._crete_date}, _etc={self._etc})"

# DAO(Data Access Object): 데이터베이스에 접근하여 조작하는 기능을 정의하는 클래스 
class DbUserDAO:
    def __init__(self, conn, file_name):
        self.conn = conn
        self.cursor = conn.cursor()
        self.file_name = file_name

    def addUser(self, vo):
        CreteDate = QDate.currentDate().toString('yyyy-MM-dd')
        self.cursor.execute("INSERT INTO user(id, school, grade, class, number, name, gender, visit_count, crete_date, etc) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                       (vo._id, vo._school, vo._grade, vo._class, vo._number, vo._name, vo._gender, vo._visit_count, CreteDate, vo._etc) )
        self.conn.commit()
        print(f'user DB ADD : {self.file_name}')
    
    def userUpdate(self, vo):
        CreteDate = QDate.currentDate().toString('yyyy-MM-dd')
        try:
            self.cursor.execute("UPDATE user SET  school=?, grade=?, class=?, number=?, name=?, gender=?, etc=? WHERE id=?",
                        ( vo._chool, vo._grade, vo._class, vo._number, vo._name, vo._gender, vo._etc, CreteDate, vo._id) )
            self.conn.commit()
            # print(f'Quiz with quiz_id={vo.quiz_id} | quiz_number={vo.quiz_number} updated.')
        except Exception as e:
            print(f'userUpdate DB ERR update :{e}')

    def userFind(self, vo):
        try:
            self.cursor.execute("SELECT * FROM user WHERE id=?", (vo._id))
            userData = self.cursor.fetchone()
            if len(self.cursor.fetchone()):
                return userData
        except Exception as e:
            print(f"[user Find ERR] {e}")



##################################################################

# print(id_dicList)

# if __name__=="__main__":
    # create_users_Table()

    # create_User("3212", 3, 2, 12 ,"홍길동", "남")
    # create_User("3211", 3, 2, 11 ,"홍길순", "여")
    # create_User("3210", 3, 2, 10 ,"이순신", "남")
    
    # a = read_User(3212)

    # read_All()